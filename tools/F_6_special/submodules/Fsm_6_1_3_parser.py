# -*- coding: utf-8 -*-
"""
Парсер табелей сотрудников.

Извлекает данные из Excel файлов табелей с поддержкой:
- Retry-логика для заблокированных файлов (OneDrive/Network)
- Динамическое определение границ данных
- Классификация строк (проекты, специальные категории, итого)
"""

import time
import warnings
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple, Any, TYPE_CHECKING, FrozenSet

from Daman_QGIS.utils import log_info, log_warning, log_error

# Подавляем предупреждения от openpyxl (совместимость с numpy 1.24+ и Python 3.12+)
# - DeprecationWarning: datetime.utcfromtimestamp() deprecated в Python 3.12
# - UserWarning: "Data Validation extension is not supported" (безопасно игнорировать)
# - FutureWarning: np.bool/np.int/np.float deprecated в numpy 1.24+
warnings.filterwarnings("ignore", category=DeprecationWarning, module="openpyxl")
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
warnings.filterwarnings("ignore", category=FutureWarning, message=".*np\\.bool.*")
warnings.filterwarnings("ignore", category=FutureWarning, message=".*np\\.int.*")
warnings.filterwarnings("ignore", category=FutureWarning, message=".*np\\.float.*")

# Ленивый импорт openpyxl для избежания ошибки np.float при загрузке плагина
# (некоторые версии openpyxl/numpy вызывают проблемы при старте)
if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet


# Специальные категории (не являются шифрами проектов)
# Используем lowercase для сравнения без учета регистра
SPECIAL_CATEGORIES = {
    "коммерческие предложения",
    "общепроизводственные вопросы",
    "отгул",
    "отпуск",
    "больничный",
    "обучение"
}

# Фиксированный порядок специальных категорий для отображения в сводном табеле
SPECIAL_CATEGORIES_ORDER = [
    "Коммерческие предложения",
    "Общепроизводственные вопросы",
    "Обучение",
    "Отгул",
    "Отпуск",
    "Больничный"
]

# Папка с табелями руководителя (фиксированный путь)
MANAGER_TIMESHEET_FOLDER = r"A:\_ГРАДОСТРОИТЕЛЬСТВО\Служебная документация\Сводные табели\Загрузка Божук"

# Базовая папка с табелями сотрудников (сетевой диск)
TIMESHEETS_BASE_FOLDER = r"\\krtsys.ru\docs\_240\Suslova\УЧЕТ РАБОЧЕГО ВРЕМЕНИ"

# Файл-шаблон со справочником шифров проектов (лист "список проектов")
PROJECT_CODES_TEMPLATE_FILE = r"\\krtsys.ru\docs\_240\Suslova\УЧЕТ РАБОЧЕГО ВРЕМЕНИ\ШАБЛОН НЕ УДАЛЯТЬ.xlsx"
PROJECT_CODES_SHEET_NAME = "список проектов"

# Названия месяцев для поиска папок
MONTH_NAMES_RU = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]

# Фиксированные ячейки шаблона
CELL_FIO = (10, 1)        # A10 - ФИО сотрудника
CELL_MONTH_START = (4, 2)  # B4 - начало месяца
CELL_MONTH_END = (4, 3)    # C4 - конец месяца
FIRST_DATA_ROW = 10        # Первая строка данных
COL_CODE = 3               # Колонка C - шифр проекта
COL_NAME = 4               # Колонка D - название объекта
COL_TOTAL = 8              # Колонка H - итого часов
COL_DAYS_START = 9         # Колонка I - начало дней (день 1)
COL_DAYS_END = 41          # Колонка AO - конец дней (день 31)


@dataclass
class ProjectRow:
    """Данные строки проекта."""
    row_number: int
    code: str
    name: Optional[str]
    total_hours: float
    daily_hours: Dict[int, float] = field(default_factory=dict)  # {день: часы}


@dataclass
class SpecialCategoryRow:
    """Данные строки специальной категории."""
    row_number: int
    category: str
    total_hours: float
    daily_hours: Dict[int, float] = field(default_factory=dict)


@dataclass
class TimesheetData:
    """Полные данные табеля сотрудника."""
    filepath: str
    filename: str
    fio: str
    month_start: Optional[date]
    month_end: Optional[date]
    year: int
    month: int
    projects: List[ProjectRow] = field(default_factory=list)
    special_categories: List[SpecialCategoryRow] = field(default_factory=list)
    itogo_row: Optional[int] = None
    total_hours: float = 0.0
    parse_errors: List[str] = field(default_factory=list)
    is_manager: bool = False

    @property
    def surname(self) -> str:
        """Извлечь фамилию из ФИО."""
        if not self.fio:
            return ""
        parts = self.fio.strip().split()
        return parts[0] if parts else ""

    @property
    def vacation_hours(self) -> float:
        """Общее количество часов отпуска."""
        for category in self.special_categories:
            if category.category.lower() == "отпуск":
                return category.total_hours
        return 0.0

    def get_vacation_hours_until_day(self, end_day: int) -> float:
        """
        Часы отпуска с 1 по указанный день включительно.

        Args:
            end_day: Последний день для расчёта (включительно)

        Returns:
            Сумма часов отпуска за период 1..end_day
        """
        for category in self.special_categories:
            if category.category.lower() == "отпуск":
                return sum(h for d, h in category.daily_hours.items() if d <= end_day)
        return 0.0

    def get_hours_until_day(self, end_day: int) -> float:
        """
        Рассчитать сумму часов с 1 по указанный день включительно.

        Args:
            end_day: Последний день для расчёта (включительно)

        Returns:
            Сумма часов за период 1..end_day
        """
        total = 0.0

        for project in self.projects:
            for day, hours in project.daily_hours.items():
                if day <= end_day:
                    total += hours

        for category in self.special_categories:
            for day, hours in category.daily_hours.items():
                if day <= end_day:
                    total += hours

        return total

    def get_filled_days_until(self, end_day: int) -> Set[int]:
        """
        Получить множество заполненных дней с 1 по указанный день.

        Args:
            end_day: Последний день для проверки (включительно)

        Returns:
            Множество дней с данными
        """
        filled: Set[int] = set()

        for project in self.projects:
            for day in project.daily_hours.keys():
                if day <= end_day:
                    filled.add(day)

        for category in self.special_categories:
            for day in category.daily_hours.keys():
                if day <= end_day:
                    filled.add(day)

        return filled


def _patch_numpy_float() -> None:
    """
    Патч для совместимости openpyxl 3.1.x с numpy 1.24+.

    NumPy 1.24 удалил np.float (deprecated с 1.20).
    Старые версии openpyxl используют np.float внутри библиотеки.
    """
    import numpy as np

    # FutureWarning уже подавлен глобально в начале модуля
    if not hasattr(np, 'float'):
        np.float = float  # type: ignore[attr-defined]
    if not hasattr(np, 'int'):
        np.int = int  # type: ignore[attr-defined]
    if not hasattr(np, 'bool'):
        np.bool = bool  # type: ignore[attr-defined]


def safe_read_excel(
    filepath: str,
    max_retries: int = 3,
    delay: float = 1.0
) -> Optional['Workbook']:
    """
    Прочитать Excel файл с retry-логикой для заблокированных файлов.

    Args:
        filepath: Путь к Excel файлу
        max_retries: Максимальное количество попыток
        delay: Задержка между попытками в секундах

    Returns:
        Workbook объект или None при ошибке
    """
    # Патч numpy для совместимости со старыми версиями openpyxl
    _patch_numpy_float()

    # Warnings уже подавлены глобально в начале модуля
    from openpyxl import load_workbook

    for attempt in range(max_retries):
        try:
            return load_workbook(filepath, read_only=False, data_only=True)
        except PermissionError:
            if attempt < max_retries - 1:
                log_warning(
                    f"Fsm_6_1_3: Файл заблокирован, попытка {attempt + 1}/{max_retries}: "
                    f"{Path(filepath).name}"
                )
                time.sleep(delay)
            else:
                log_error(
                    f"Fsm_6_1_3: Не удалось прочитать файл (заблокирован): "
                    f"{Path(filepath).name}"
                )
                return None
        except FileNotFoundError:
            log_error(f"Fsm_6_1_3: Файл не найден: {filepath}")
            return None
        except Exception as e:
            log_error(f"Fsm_6_1_3: Ошибка чтения файла {Path(filepath).name}: {e}")
            return None
    return None


def find_data_bounds(ws: 'Worksheet') -> Tuple[int, int, Optional[int], int]:
    """
    Найти границы данных в табеле.

    Returns:
        Tuple (first_data_row, last_data_row, itogo_row, special_categories_end_row)
        itogo_row может быть None если строка "Итого" не найдена
        special_categories_end_row - последняя строка для поиска спец. категорий
    """
    first_row = FIRST_DATA_ROW
    last_data_row = FIRST_DATA_ROW
    itogo_row = None
    special_categories_end_row = FIRST_DATA_ROW

    # Ограничиваем поиск разумным количеством строк
    max_search_rows = min(ws.max_row + 1, 100)

    for row in range(FIRST_DATA_ROW, max_search_rows):
        # Проверяем строку "Итого" в колонках E-H (5-8)
        # В шаблоне "Итого" находится в колонке G или H
        for col in range(5, 9):
            cell_val = ws.cell(row, col).value
            if cell_val and "итого" in str(cell_val).lower():
                if itogo_row is None:
                    itogo_row = row
                break

        if itogo_row and row > itogo_row:
            # После Итого проверяем специальные категории в колонке D
            cell_d = ws.cell(row, COL_NAME).value
            if cell_d and is_special_category(cell_d):
                special_categories_end_row = row
            continue

        # До Итого - ищем строки с данными (шифры в колонке C)
        cell_c = ws.cell(row, COL_CODE).value
        if cell_c:
            cell_c_str = str(cell_c).strip()
            if cell_c_str and not "итого" in cell_c_str.lower():
                last_data_row = row

    return first_row, last_data_row, itogo_row, special_categories_end_row


def is_project_row(cell_value: Any) -> bool:
    """
    Проверить, является ли строка проектом (не специальной категорией).

    Args:
        cell_value: Значение ячейки колонки C (шифр)

    Returns:
        True если это строка проекта
    """
    if not cell_value:
        return False

    cell_str = str(cell_value).strip()
    if not cell_str:
        return False

    cell_lower = cell_str.lower()

    # Не проект если это специальная категория или итого
    if cell_lower in SPECIAL_CATEGORIES:
        return False
    if "итого" in cell_lower:
        return False

    return True


def is_special_category(cell_value: Any) -> bool:
    """
    Проверить, является ли строка специальной категорией.

    Args:
        cell_value: Значение ячейки (колонка C или D)

    Returns:
        True если это специальная категория
    """
    if not cell_value:
        return False

    cell_str = str(cell_value).strip().lower()
    return cell_str in SPECIAL_CATEGORIES


def get_special_category_from_row(ws: 'Worksheet', row: int) -> Optional[str]:
    """
    Получить специальную категорию из строки.

    Специальные категории могут быть в колонке C или D.

    Args:
        ws: Worksheet
        row: Номер строки

    Returns:
        Название категории или None
    """
    # Сначала проверяем колонку C
    cell_c = ws.cell(row, COL_CODE).value
    if cell_c and is_special_category(cell_c):
        return str(cell_c).strip()

    # Затем проверяем колонку D (название)
    cell_d = ws.cell(row, COL_NAME).value
    if cell_d and is_special_category(cell_d):
        return str(cell_d).strip()

    return None


def _parse_daily_hours(ws: 'Worksheet', row: int, days_in_month: int) -> Dict[int, float]:
    """
    Извлечь часы по дням из строки.

    Args:
        ws: Worksheet
        row: Номер строки
        days_in_month: Количество дней в месяце

    Returns:
        Словарь {день: часы}
    """
    daily_hours = {}

    for day in range(1, days_in_month + 1):
        col = COL_DAYS_START + day - 1
        if col > COL_DAYS_END:
            break

        value = ws.cell(row, col).value
        if value is not None:
            try:
                hours = float(value)
                if hours > 0:
                    daily_hours[day] = hours
            except (ValueError, TypeError):
                pass

    return daily_hours


def _is_excel_error(value) -> bool:
    """
    Проверить, является ли значение ошибкой Excel (#N/A, #VALUE!, #Н/Д и т.д.).

    openpyxl с data_only=True возвращает кэшированные ошибки формул как строки.

    Args:
        value: Значение ячейки

    Returns:
        True если это ошибка Excel
    """
    if not isinstance(value, str):
        return False
    v = value.strip()
    # Все ошибки Excel начинаются с '#': #N/A, #VALUE!, #REF!, #NAME?, #NUM!, #NULL!, #DIV/0!
    # Русские варианты: #Н/Д, #ЗНАЧ!, #ССЫЛКА!, #ИМЯ?, #ЧИСЛО!, #ПУСТО!, #ДЕЛ/0!
    return v.startswith('#') and len(v) >= 3


def normalize_project_code(code: str) -> str:
    """
    Нормализовать шифр проекта к верхнему регистру.

    Сотрудники могут вводить шифры в разном регистре (25-п-1, 25-П-1),
    эта функция приводит их к единому виду (25-П-1).

    Args:
        code: Исходный шифр проекта

    Returns:
        Шифр в верхнему регистре
    """
    return code.upper() if code else ""


def _parse_row(ws: 'Worksheet', row: int, days_in_month: int) -> Tuple[str, Optional[str], float, Dict[int, float]]:
    """
    Извлечь данные из строки табеля.

    Фильтрует ошибки Excel (#N/A, #Н/Д и т.д.) в шифре и названии.

    Returns:
        Tuple (code, name, total_hours, daily_hours)
    """
    code = ws.cell(row, COL_CODE).value
    if _is_excel_error(code):
        log_warning(f"Fsm_6_1_3: Ошибка Excel в шифре (строка {row}): '{code}' -> пропущено")
        code = None
    code_str = str(code).strip() if code else ""
    # Нормализуем шифр к верхнему регистру
    code_str = normalize_project_code(code_str)

    name = ws.cell(row, COL_NAME).value
    if _is_excel_error(name):
        log_warning(f"Fsm_6_1_3: Ошибка Excel в названии (строка {row}, шифр '{code_str}'): '{name}' -> пропущено")
        name = None
    name_str = str(name).strip() if name else None

    total_value = ws.cell(row, COL_TOTAL).value
    total_hours = 0.0
    if total_value is not None:
        try:
            total_hours = float(total_value)
        except (ValueError, TypeError):
            pass

    daily_hours = _parse_daily_hours(ws, row, days_in_month)

    return code_str, name_str, total_hours, daily_hours


def _get_days_in_month(year: int, month: int) -> int:
    """Получить количество дней в месяце."""
    import calendar
    return calendar.monthrange(year, month)[1]


def parse_timesheet(filepath: str) -> Optional[TimesheetData]:
    """
    Распарсить табель сотрудника.

    Args:
        filepath: Путь к Excel файлу табеля

    Returns:
        TimesheetData или None при критической ошибке чтения
    """
    path = Path(filepath)
    filename = path.name

    log_info(f"Fsm_6_1_3: Парсинг файла {filename}")

    # Открываем файл с retry-логикой
    # Warnings уже подавлены глобально в начале модуля
    wb = safe_read_excel(filepath)
    if wb is None:
        return None

    try:
        # Получаем лист "табель"
        sheet_names = wb.sheetnames
        if "табель" in sheet_names:
            ws = wb["табель"]
        elif len(sheet_names) > 0:
            ws = wb[sheet_names[0]]
            log_warning(f"Fsm_6_1_3: Лист 'табель' не найден, используется '{sheet_names[0]}'")
        else:
            log_error(f"Fsm_6_1_3: Файл не содержит листов: {filename}")
            return None

        # Извлекаем метаданные
        fio = ws.cell(*CELL_FIO).value
        fio_str = str(fio).strip() if fio else ""

        month_start_value = ws.cell(*CELL_MONTH_START).value
        month_end_value = ws.cell(*CELL_MONTH_END).value

        # Парсим даты
        month_start: Optional[date] = None
        month_end: Optional[date] = None
        year = datetime.now().year
        month = datetime.now().month

        if isinstance(month_start_value, datetime):
            month_start = month_start_value.date()
            year = month_start.year
            month = month_start.month
        elif isinstance(month_start_value, date):
            month_start = month_start_value
            year = month_start.year
            month = month_start.month

        if isinstance(month_end_value, datetime):
            month_end = month_end_value.date()
        elif isinstance(month_end_value, date):
            month_end = month_end_value

        days_in_month = _get_days_in_month(year, month)

        # Создаем объект результата
        result = TimesheetData(
            filepath=filepath,
            filename=filename,
            fio=fio_str,
            month_start=month_start,
            month_end=month_end,
            year=year,
            month=month
        )

        # Находим границы данных
        first_row, last_row, itogo_row, special_end_row = find_data_bounds(ws)
        result.itogo_row = itogo_row

        # Определяем конечную строку для парсинга проектов (до Итого)
        projects_end_row = itogo_row if itogo_row else last_row + 1

        # Парсим все строки до Итого (проекты и специальные категории)
        for row in range(first_row, projects_end_row):
            code = ws.cell(row, COL_CODE).value
            name = ws.cell(row, COL_NAME).value

            # Проверяем проект (шифр в колонке C)
            if is_project_row(code):
                code_str, name_str, total_hours, daily_hours = _parse_row(ws, row, days_in_month)

                # Пропускаем пустые строки проектов (нулевые часы)
                if total_hours > 0 or daily_hours:
                    project = ProjectRow(
                        row_number=row,
                        code=code_str,
                        name=name_str,
                        total_hours=total_hours,
                        daily_hours=daily_hours
                    )
                    result.projects.append(project)
                    result.total_hours += total_hours

            # Проверяем специальные категории (в колонке D, когда колонка C пустая)
            elif name and is_special_category(name):
                category_name = str(name).strip()
                _, _, total_hours, daily_hours = _parse_row(ws, row, days_in_month)

                # Сохраняем специальные категории с данными
                if total_hours > 0 or daily_hours:
                    special = SpecialCategoryRow(
                        row_number=row,
                        category=category_name,
                        total_hours=total_hours,
                        daily_hours=daily_hours
                    )
                    result.special_categories.append(special)
                    result.total_hours += total_hours
                    log_info(f"Fsm_6_1_3: Найдена спец.категория '{category_name}' на строке {row}, часов={total_hours}")

        log_info(
            f"Fsm_6_1_3: Распарсен {filename}: "
            f"ФИО={fio_str}, проектов={len(result.projects)}, "
            f"спец.категорий={len(result.special_categories)}, "
            f"всего часов={result.total_hours}"
        )

        return result

    except Exception as e:
        log_error(f"Fsm_6_1_3: Ошибка парсинга {filename}: {e}")
        return None

    finally:
        wb.close()


def parse_timesheets_from_folder(
    folder_path: str,
    valid_surnames: Optional[Set[str]] = None,
    target_month: Optional[int] = None
) -> List[TimesheetData]:
    """
    Распарсить табели из папки с фильтрацией по фамилиям и месяцу.

    Файлы должны соответствовать паттерну Фамилия_М.xlsx или Фамилия_ММ.xlsx,
    где фамилия есть в базе сотрудников и месяц соответствует целевому.

    Args:
        folder_path: Путь к папке с табелями
        valid_surnames: Множество допустимых фамилий (lowercase).
                       Если None - читаем все файлы без фильтрации.
        target_month: Целевой месяц (1-12). Если None - текущий месяц.

    Returns:
        Список успешно распарсенных табелей
    """
    import re
    from datetime import datetime

    folder = Path(folder_path)
    if not folder.exists():
        log_error(f"Fsm_6_1_3: Папка не существует: {folder_path}")
        return []

    if not folder.is_dir():
        log_error(f"Fsm_6_1_3: Путь не является папкой: {folder_path}")
        return []

    # Определяем целевой месяц
    if target_month is None:
        target_month = datetime.now().month
        log_info(f"Fsm_6_1_3: target_month не передан, используется текущий: {target_month}")
    else:
        log_info(f"Fsm_6_1_3: Переданный target_month: {target_month}")

    # Паттерн имени файла: Фамилия_М.xlsx или Фамилия_ММ.xlsx
    filename_pattern = re.compile(r'^([А-Яа-яЁё]+)_(\d{1,2})\.xlsx$', re.UNICODE)

    # Находим все xlsx файлы
    xlsx_files = list(folder.glob("*.xlsx"))

    # Исключаем временные файлы Excel (начинаются с ~$)
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

    if not xlsx_files:
        log_warning(f"Fsm_6_1_3: Не найдено .xlsx файлов в папке: {folder_path}")
        return []

    log_info(f"Fsm_6_1_3: Найдено {len(xlsx_files)} xlsx файлов в папке")

    # Фильтруем файлы по паттерну, фамилии и месяцу
    matched_files: List[Path] = []
    skipped_files: List[str] = []

    for xlsx_file in xlsx_files:
        filename = xlsx_file.name
        match = filename_pattern.match(filename)

        if not match:
            skipped_files.append(f"{filename} (неверный формат имени)")
            continue

        surname = match.group(1)
        month_str = match.group(2)

        # Проверяем месяц
        try:
            file_month = int(month_str)
            if file_month != target_month:
                skipped_files.append(f"{filename} (месяц {file_month}, ожидается {target_month})")
                continue
        except ValueError:
            skipped_files.append(f"{filename} (неверный месяц)")
            continue

        # Проверяем фамилию в базе (если указан фильтр)
        if valid_surnames is not None:
            if surname.lower() not in valid_surnames:
                skipped_files.append(f"{filename} (фамилия не в базе)")
                continue

        matched_files.append(xlsx_file)

    # Логируем пропущенные файлы
    if skipped_files:
        log_info(f"Fsm_6_1_3: Пропущено {len(skipped_files)} файлов:")
        for skipped in skipped_files[:10]:  # Показываем первые 10
            log_info(f"  - {skipped}")
        if len(skipped_files) > 10:
            log_info(f"  ... и еще {len(skipped_files) - 10}")

    if not matched_files:
        log_warning(f"Fsm_6_1_3: Нет подходящих файлов табелей для месяца {target_month}")
        return []

    log_info(f"Fsm_6_1_3: Найдено {len(matched_files)} подходящих файлов для обработки")

    # Парсим найденные файлы
    results: List[TimesheetData] = []
    for xlsx_file in matched_files:
        timesheet = parse_timesheet(str(xlsx_file))
        if timesheet is not None:
            results.append(timesheet)

    log_info(
        f"Fsm_6_1_3: Успешно распарсено {len(results)} из {len(matched_files)} файлов"
    )

    return results


def get_manager_timesheet(
    valid_surnames: Optional[Set[str]] = None,
    target_month: Optional[int] = None
) -> Optional[TimesheetData]:
    """
    Получить табель руководителя из фиксированной папки.

    Табели руководителя хранятся в отдельной папке MANAGER_TIMESHEET_FOLDER.
    Функция находит и парсит табель за указанный месяц.

    Args:
        valid_surnames: Множество допустимых фамилий (lowercase).
                       Если None - читаем без фильтрации по фамилии.
        target_month: Целевой месяц (1-12). Если None - текущий месяц.

    Returns:
        TimesheetData если табель найден и успешно распарсен, иначе None
    """
    import re
    from datetime import datetime

    folder = Path(MANAGER_TIMESHEET_FOLDER)

    # Проверяем доступность папки (может быть на сетевом диске)
    if not folder.exists():
        log_info(f"Fsm_6_1_3: Папка руководителя недоступна: {MANAGER_TIMESHEET_FOLDER}")
        return None

    # Определяем целевой месяц
    if target_month is None:
        target_month = datetime.now().month

    # Паттерн имени файла: Фамилия_М.xlsx или Фамилия_ММ.xlsx
    filename_pattern = re.compile(r'^([А-Яа-яЁё]+)_(\d{1,2})\.xlsx$', re.UNICODE)

    # Находим все xlsx файлы
    xlsx_files = list(folder.glob("*.xlsx"))

    # Исключаем временные файлы Excel (начинаются с ~$)
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

    if not xlsx_files:
        log_info(f"Fsm_6_1_3: Не найдено файлов в папке руководителя")
        return None

    # Ищем файл за нужный месяц
    for xlsx_file in xlsx_files:
        filename = xlsx_file.name
        match = filename_pattern.match(filename)

        if not match:
            continue

        surname = match.group(1)
        month_str = match.group(2)

        # Проверяем месяц
        try:
            file_month = int(month_str)
            if file_month != target_month:
                continue
        except ValueError:
            continue

        # Проверяем фамилию в базе (если указан фильтр)
        if valid_surnames is not None:
            if surname.lower() not in valid_surnames:
                log_warning(
                    f"Fsm_6_1_3: Табель руководителя {filename} - "
                    f"фамилия '{surname}' не найдена в базе сотрудников"
                )
                continue

        # Нашли подходящий файл - парсим
        log_info(f"Fsm_6_1_3: Найден табель руководителя: {filename}")
        timesheet = parse_timesheet(str(xlsx_file))

        if timesheet is not None:
            timesheet.is_manager = True
            log_info(f"Fsm_6_1_3: Табель руководителя успешно распарсен: {timesheet.fio}")
            return timesheet
        else:
            log_warning(f"Fsm_6_1_3: Не удалось распарсить табель руководителя: {filename}")

    log_info(f"Fsm_6_1_3: Табель руководителя за месяц {target_month} не найден")
    return None


def find_timesheet_folder(year: int, month: int) -> Optional[str]:
    """
    Найти папку с табелями сотрудников по году и месяцу.

    Ищет папку в структуре:
    TIMESHEETS_BASE_FOLDER / {год} / {месяц}_{Название}_{год}

    Поддерживает форматы имени папки месяца:
    - 2_Февраль_2026 (без ведущего нуля)
    - 02_Февраль_2026 (с ведущим нулём)

    Args:
        year: Год (например, 2026)
        month: Месяц (1-12)

    Returns:
        Путь к папке или None если не найдена
    """
    base_folder = Path(TIMESHEETS_BASE_FOLDER)

    # Проверяем доступность базовой папки
    if not base_folder.exists():
        log_warning(f"Fsm_6_1_3: Базовая папка табелей недоступна: {TIMESHEETS_BASE_FOLDER}")
        return None

    # Папка года
    year_folder = base_folder / str(year)
    if not year_folder.exists():
        log_warning(f"Fsm_6_1_3: Папка года не найдена: {year_folder}")
        return None

    # Название месяца
    month_name = MONTH_NAMES_RU[month]

    # Варианты имени папки месяца (с ведущим нулём и без)
    folder_variants = [
        f"{month}_{month_name}_{year}",      # 2_Февраль_2026
        f"{month:02d}_{month_name}_{year}",  # 02_Февраль_2026
    ]

    # Ищем подходящую папку
    for folder_name in folder_variants:
        month_folder = year_folder / folder_name
        if month_folder.exists() and month_folder.is_dir():
            log_info(f"Fsm_6_1_3: Найдена папка табелей: {month_folder}")
            return str(month_folder)

    # Если точное совпадение не найдено, ищем по паттерну
    # (на случай если название месяца немного отличается)
    import re
    pattern = re.compile(rf'^0?{month}_[А-Яа-яЁё]+_{year}$', re.UNICODE)

    for item in year_folder.iterdir():
        if item.is_dir() and pattern.match(item.name):
            log_info(f"Fsm_6_1_3: Найдена папка табелей (по паттерну): {item}")
            return str(item)

    log_warning(
        f"Fsm_6_1_3: Папка табелей за {month:02d}.{year} не найдена. "
        f"Ожидались варианты: {folder_variants}"
    )
    return None


def extract_project_codes_from_sheet(
    filepath: str = PROJECT_CODES_TEMPLATE_FILE,
    sheet_name: str = PROJECT_CODES_SHEET_NAME
) -> List[Dict[str, str]]:
    """
    Извлечь шифры проектов из листа справочника.

    По умолчанию загружает из сетевого шаблона ШАБЛОН НЕ УДАЛЯТЬ.xlsx,
    лист "список проектов".

    Колонки:
    - A: Код объекта
    - B: Объект (рабочее название)
    - C: Номер договора
    - D: Дата договора
    - E: Заказчик

    Args:
        filepath: Путь к Excel файлу (по умолчанию PROJECT_CODES_TEMPLATE_FILE)
        sheet_name: Имя листа со справочником (по умолчанию PROJECT_CODES_SHEET_NAME)

    Returns:
        Список словарей с данными проектов
    """
    # Проверяем доступность файла
    from pathlib import Path
    if not Path(filepath).exists():
        log_warning(f"Fsm_6_1_3: Файл справочника недоступен: {filepath}")
        return []

    wb = safe_read_excel(filepath)
    if wb is None:
        return []

    try:
        if sheet_name not in wb.sheetnames:
            log_warning(f"Fsm_6_1_3: Лист '{sheet_name}' не найден в {filepath}")
            return []

        ws = wb[sheet_name]
        projects = []

        # Пропускаем заголовок (строка 1)
        for row in range(2, ws.max_row + 1):
            code = ws.cell(row, 1).value  # A - Код объекта
            name = ws.cell(row, 2).value  # B - Объект (рабочее название)
            contract_number = ws.cell(row, 3).value  # C - Номер договора
            contract_date = ws.cell(row, 4).value  # D - Дата договора
            customer = ws.cell(row, 5).value  # E - Заказчик

            if code:
                project = {
                    "code": str(code).strip().upper(),  # Нормализуем к верхнему регистру
                    "name": str(name).strip() if name else "",
                    "contract_number": str(contract_number).strip() if contract_number else "",
                    "contract_date": str(contract_date).strip() if contract_date else "",
                    "customer": str(customer).strip() if customer else ""
                }
                projects.append(project)

        log_info(f"Fsm_6_1_3: Загружено {len(projects)} шифров проектов из справочника")
        return projects

    finally:
        wb.close()


def load_valid_project_codes() -> Set[str]:
    """
    Загрузить множество валидных шифров проектов из сетевого справочника.

    Оптимизировано: читает только колонку A через iter_rows (streaming).
    Это значительно быстрее чем extract_project_codes_from_sheet(),
    которая читает все 5 колонок поячеечно.

    Returns:
        Множество шифров проектов (в верхнем регистре)
    """
    filepath = PROJECT_CODES_TEMPLATE_FILE
    sheet_name = PROJECT_CODES_SHEET_NAME

    if not Path(filepath).exists():
        log_warning(f"Fsm_6_1_3: Файл справочника недоступен: {filepath}")
        return set()

    start_time = time.time()
    wb = safe_read_excel(filepath)
    if wb is None:
        return set()

    try:
        if sheet_name not in wb.sheetnames:
            log_warning(f"Fsm_6_1_3: Лист '{sheet_name}' не найден в {filepath}")
            return set()

        ws = wb[sheet_name]
        codes: Set[str] = set()

        # Читаем только колонку A через iter_rows (streaming в read_only)
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            code = row[0]
            if code:
                codes.add(str(code).strip().upper())

        elapsed = time.time() - start_time
        log_info(f"Fsm_6_1_3: Загружено {len(codes)} уникальных шифров "
                f"для валидации за {elapsed:.1f}с")
        return codes

    finally:
        wb.close()
