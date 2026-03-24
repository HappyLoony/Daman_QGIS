# -*- coding: utf-8 -*-
"""
Генератор сводного табеля (Сводный_табель.xlsx).

Создает матрицу: сотрудники x шифры проектов с суммарными часами.
"""

from datetime import datetime, date, timedelta
from pathlib import Path
from typing import List, Optional, Dict, Set, Tuple, TYPE_CHECKING

from Daman_QGIS.utils import log_info, log_warning, log_error
from Daman_QGIS.managers.reference.submodules import ProductionCalendarManager
from .Fsm_6_1_3_parser import TimesheetData, SPECIAL_CATEGORIES, SPECIAL_CATEGORIES_ORDER

# Ленивый импорт openpyxl для избежания ошибки np.float при загрузке плагина
if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Шрифт по умолчанию для генерируемых файлов
DEFAULT_FONT_NAME = "Times New Roman"
DEFAULT_FONT_SIZE = 11


class SummaryTimesheetGenerator:
    """Генератор сводного табеля."""

    # Имя выходного файла
    OUTPUT_FILENAME = "Сводный_табель.xlsx"

    # Фиксированные ширины колонок для специальных категорий
    CATEGORY_COLUMN_WIDTHS: Dict[str, int] = {
        "Коммерческие предложения": 16,
        "Общепроизводственные вопросы": 14,
        "Обучение": 11,
        "Отгул": 7,
        "Отпуск": 8,
        "Больничный": 14,
    }

    # Ширины служебных колонок
    TOTAL_COL_WIDTH = 10
    DEVIATION_COL_WIDTH = 13

    def __init__(self):
        self._styles_initialized = False

        # Стили (инициализируются лениво)
        self._stripe_fill_5: Optional['PatternFill'] = None
        self._stripe_fill_15: Optional['PatternFill'] = None
        self._thin_side: Optional['Side'] = None
        self._thick_side: Optional['Side'] = None
        self._thin_border: Optional['Border'] = None
        self._thick_border: Optional['Border'] = None

    def _init_styles(self) -> None:
        """Инициализировать стили openpyxl (ленивая загрузка)."""
        if self._styles_initialized:
            return

        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        # Стили заливок
        # "Белый, фон 1, более темный оттенок 5%" = D9D9D9 (светло-серый)
        # "Белый, фон 1, более темный оттенок 15%" = BFBFBF (серый)
        self._stripe_fill_5 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        self._stripe_fill_15 = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

        # Границы
        self._thin_side = Side(style='thin')
        self._thick_side = Side(style='medium')

        self._thin_border = Border(
            left=self._thin_side,
            right=self._thin_side,
            top=self._thin_side,
            bottom=self._thin_side
        )

        self._thick_border = Border(
            left=self._thick_side,
            right=self._thick_side,
            top=self._thick_side,
            bottom=self._thick_side
        )

        self._styles_initialized = True

    def _apply_outer_thick_border(self, ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
        """
        Применить толстые внешние границы к диапазону ячеек.

        Args:
            ws: Worksheet
            start_row: Начальная строка
            start_col: Начальная колонка
            end_row: Конечная строка
            end_col: Конечная колонка
        """
        from openpyxl.styles import Border
        self._init_styles()

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row, col)

                # Определяем какие стороны должны быть толстыми
                left = self._thick_side if col == start_col else self._thin_side
                right = self._thick_side if col == end_col else self._thin_side
                top = self._thick_side if row == start_row else self._thin_side
                bottom = self._thick_side if row == end_row else self._thin_side

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def _collect_unique_codes(self, timesheets: List[TimesheetData]) -> Tuple[List[str], List[str]]:
        """
        Собрать уникальные шифры проектов из всех табелей.

        Специальные категории всегда возвращаются полным списком в фиксированном порядке.

        Args:
            timesheets: Список табелей

        Returns:
            Tuple (шифры_проектов, специальные_категории)
            - шифры_проектов: отсортированы лексикографически
            - специальные_категории: фиксированный порядок из SPECIAL_CATEGORIES_ORDER
        """
        codes: Set[str] = set()

        for ts in timesheets:
            # Шифры проектов
            for project in ts.projects:
                if project.code and project.code.lower() not in SPECIAL_CATEGORIES:
                    codes.add(project.code)

        # Шифры сортируем лексикографически
        # Категории всегда возвращаем полным списком в фиксированном порядке
        return sorted(codes), SPECIAL_CATEGORIES_ORDER.copy()

    def _collect_project_names(self, timesheets: List[TimesheetData]) -> Dict[str, str]:
        """
        Собрать названия проектов по шифрам.

        Args:
            timesheets: Список табелей

        Returns:
            Словарь {шифр: название}
        """
        names: Dict[str, str] = {}

        for ts in timesheets:
            for project in ts.projects:
                if project.code and project.name:
                    if project.code not in names:
                        names[project.code] = project.name
                        log_info(
                            f"Fsm_6_1_5: Шифр '{project.code}' -> "
                            f"название '{project.name}' (из {ts.filename})"
                        )

        return names

    def _build_data_matrix(
        self,
        timesheets: List[TimesheetData],
        codes: List[str],
        categories: List[str],
        end_day: Optional[int] = None
    ) -> Tuple[List[Tuple[str, Dict[str, float]]], Dict[str, float]]:
        """
        Построить матрицу данных.

        Args:
            timesheets: Список табелей
            codes: Список шифров проектов (колонки)
            categories: Список специальных категорий (колонки после шифров)
            end_day: Последний день для расчёта часов (включительно).
                     Если None - используются все дни.

        Returns:
            Tuple (данные_сотрудников, суммы_по_колонкам)
            данные_сотрудников: [(фамилия, {шифр_или_категория: часы}), ...]
            суммы_по_колонкам: {шифр_или_категория: общая_сумма}
        """
        # Объединяем шифры и категории для инициализации totals
        all_columns = codes + categories

        employees_data: List[Tuple[str, Dict[str, float]]] = []
        column_totals: Dict[str, float] = {col: 0.0 for col in all_columns}

        # Создаём маппинг lowercase -> canonical для категорий
        category_mapping: Dict[str, str] = {cat.lower(): cat for cat in categories}

        def sum_hours_until(daily_hours: Dict[int, float], limit_day: Optional[int]) -> float:
            """Сумма часов до указанного дня включительно."""
            if limit_day is None:
                return sum(daily_hours.values())
            return sum(h for d, h in daily_hours.items() if d <= limit_day)

        for ts in timesheets:
            # Собираем часы по шифрам и категориям для сотрудника
            hours_by_column: Dict[str, float] = {}

            # Проекты
            for project in ts.projects:
                code = project.code
                if code and code in column_totals:
                    # Считаем часы только до end_day
                    hours = sum_hours_until(project.daily_hours, end_day)
                    if hours > 0:
                        if code in hours_by_column:
                            hours_by_column[code] += hours
                        else:
                            hours_by_column[code] = hours
                        column_totals[code] += hours

            # Специальные категории (сопоставляем без учёта регистра)
            for category in ts.special_categories:
                cat_name = category.category
                if cat_name:
                    # Находим каноническое имя категории
                    canonical_name = category_mapping.get(cat_name.lower())
                    if canonical_name:
                        # Считаем часы только до end_day
                        hours = sum_hours_until(category.daily_hours, end_day)
                        if hours > 0:
                            if canonical_name in hours_by_column:
                                hours_by_column[canonical_name] += hours
                            else:
                                hours_by_column[canonical_name] = hours
                            column_totals[canonical_name] += hours

            employees_data.append((ts.fio, hours_by_column))

        return employees_data, column_totals

    def generate(
        self,
        timesheets: List[TimesheetData],
        output_folder: str,
        report_end_date: Optional[date] = None,
        employee_rates: Optional[Dict[str, float]] = None
    ) -> Optional[str]:
        """
        Сгенерировать сводный табель.

        Args:
            timesheets: Список данных табелей
            output_folder: Папка для сохранения
            report_end_date: Дата окончания расчетного периода (включительно).
                           Если None - используется вчерашняя дата.
            employee_rates: Словарь {ФИО: ставка} для расчёта индивидуальных норм.
                           Если None или ФИО отсутствует - используется ставка 1.0.

        Returns:
            Путь к созданному файлу или None при ошибке
        """
        # Патч numpy для совместимости со старыми версиями openpyxl
        from .Fsm_6_1_3_parser import _patch_numpy_float
        _patch_numpy_float()

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        if not timesheets:
            log_warning("Fsm_6_1_5: Нет табелей для сводки")
            return None

        log_info(f"Fsm_6_1_5: Генерация сводного табеля для {len(timesheets)} сотрудников")

        # Инициализируем стили
        self._init_styles()

        try:
            # Сортируем по фамилии
            timesheets_sorted = sorted(timesheets, key=lambda t: t.surname.lower())

            # Собираем шифры проектов и специальные категории
            codes, categories = self._collect_unique_codes(timesheets_sorted)
            names = self._collect_project_names(timesheets_sorted)

            # Объединяем для колонок: сначала шифры, потом категории
            all_columns = codes + categories

            if not all_columns:
                log_warning("Fsm_6_1_5: Не найдено шифров проектов и категорий")
                return None

            # Строим матрицу данных (с учётом end_day)
            end_day = report_end_date.day if report_end_date else None
            employees_data, column_totals = self._build_data_matrix(
                timesheets_sorted, codes, categories, end_day=end_day
            )

            # Определяем период
            first_ts = timesheets_sorted[0]
            month_start = first_ts.month_start
            year = first_ts.year
            month = first_ts.month

            # Конечная дата из параметра или вчерашняя дата по умолчанию
            if report_end_date is None:
                report_end_date = date.today() - timedelta(days=1)

            log_info(f"Fsm_6_1_5: Расчетный период: {month_start} - {report_end_date}")

            # Рассчитываем нормы часов с учётом предпраздничных дней
            calendar_manager = ProductionCalendarManager()

            # Норма за период (от начала месяца до вчера включительно)
            if month_start:
                norm_period = calendar_manager.get_work_hours_for_period(
                    month_start, report_end_date
                )
            else:
                norm_period = 0.0

            # Норма за весь месяц
            norm_month = calendar_manager.get_work_hours_for_month(year, month)

            log_info(
                f"Fsm_6_1_5: Норма часов - за период: {norm_period}, за месяц: {norm_month}"
            )

            # Создаем книгу
            wb = Workbook()

            # Устанавливаем дефолтный шрифт Workbook (влияет на единицы ширины колонок)
            normal_style = wb._named_styles['Normal']
            normal_style.font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)

            ws = wb.active
            ws.title = "Сводка"

            # Шрифты (все размером 11)
            default_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
            bold_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)
            italic_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, italic=True)
            title_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)

            # Колонки
            total_col = len(all_columns) + 2
            deviation_col = total_col + 1
            last_data_col = len(all_columns) + 1  # Последняя колонка с данными (до Итого)

            # === Строка 1: Заголовок ===
            ws.cell(1, 1).value = "План-график загрузки отдела на период"
            ws.cell(1, 1).font = title_font
            ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # D1:E1 - даты с толстыми границами
            if month_start:
                ws.cell(1, 4).value = month_start.strftime("%d.%m.%Y")
                ws.cell(1, 4).font = bold_font
                ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(1, 4).border = self._thick_border

            ws.cell(1, 5).value = report_end_date.strftime("%d.%m.%Y")
            ws.cell(1, 5).font = bold_font
            ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(1, 5).border = self._thick_border

            # === Строка 4: Заголовки колонок ===
            ws.cell(4, 1).value = "ФИО сотрудника"
            ws.cell(4, 1).font = bold_font
            ws.cell(4, 1).border = self._thin_border
            ws.cell(4, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.cell(4, 2).value = "Шифр объекта"
            ws.cell(4, 2).font = bold_font
            ws.cell(4, 2).border = self._thin_border
            ws.cell(4, 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # === Строки 4-5: Шифры проектов (строка 5) и специальные категории (строка 4, объединяется с 5) ===
            ws.cell(5, 1).value = ""
            ws.cell(5, 1).border = self._thin_border

            # Определяем границу между шифрами и категориями
            first_category_col = 1 + len(codes) + 1  # Первая колонка с категорией

            for col_idx, column in enumerate(all_columns, start=2):
                if col_idx < first_category_col:
                    # Шифр проекта - в строку 5
                    cell = ws.cell(5, col_idx)
                    cell.value = column
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = self._thin_border
                    # Чередующаяся заливка: нечётные колонки (B=2, D=4...) с заливкой 5%
                    if col_idx % 2 == 0:  # B=2, D=4, F=6...
                        cell.fill = self._stripe_fill_5
                else:
                    # Специальная категория - в строку 4 (будет объединена с 5)
                    cell = ws.cell(4, col_idx)
                    cell.value = column
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = self._thin_border

            # === Нормы часов в верхних строках ===
            # Колонка перед total_col (M) - подписи, колонка total_col (N) - значения
            label_col = total_col - 1

            # Строка 1: норма за период
            ws.cell(1, label_col).value = "Норма часов за период"
            ws.cell(1, label_col).font = italic_font
            ws.cell(1, label_col).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

            ws.cell(1, total_col).value = norm_period
            ws.cell(1, total_col).font = bold_font
            ws.cell(1, total_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Строка 2: норма за месяц
            ws.cell(2, label_col).value = "Норма часов за месяц"
            ws.cell(2, label_col).font = italic_font
            ws.cell(2, label_col).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

            ws.cell(2, total_col).value = norm_month
            ws.cell(2, total_col).font = bold_font
            ws.cell(2, total_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Заголовок "Итого часов за период" в строке 4 (будет объединено с 5)
            ws.cell(4, total_col).value = "Итого часов за период"
            ws.cell(4, total_col).font = bold_font
            ws.cell(4, total_col).border = self._thin_border
            ws.cell(4, total_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Заголовок "Отклонение от нормы" в строке 4 (будет объединено с 5)
            ws.cell(4, deviation_col).value = "Отклонение от нормы"
            ws.cell(4, deviation_col).font = bold_font
            ws.cell(4, deviation_col).border = self._thin_border
            ws.cell(4, deviation_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # === Строки 6+: Данные сотрудников ===
            current_row = 6
            first_data_row = 6  # Запоминаем первую строку данных
            grand_total = 0.0
            total_expected_hours = 0.0  # Сумма индивидуальных норм для общего отклонения

            for fio, hours_by_column in employees_data:
                # Получаем ставку сотрудника
                rate = 1.0
                if employee_rates and fio in employee_rates:
                    rate = employee_rates[fio]

                # Отпуск считается по полной ставке (1.0), остальное по rate
                vacation_hours = hours_by_column.get("Отпуск", 0.0)
                if rate == 1.0:
                    employee_norm = norm_period
                else:
                    work_norm = norm_period - vacation_hours
                    employee_norm = work_norm * rate + vacation_hours
                total_expected_hours += employee_norm

                # ФИО
                ws.cell(current_row, 1).value = fio
                ws.cell(current_row, 1).font = default_font
                ws.cell(current_row, 1).border = self._thin_border
                ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Часы по шифрам и категориям
                employee_total = 0.0
                for col_idx, column in enumerate(all_columns, start=2):
                    cell = ws.cell(current_row, col_idx)
                    hours = hours_by_column.get(column, 0)
                    if hours > 0:
                        cell.value = hours
                        employee_total += hours
                    cell.font = default_font
                    cell.border = self._thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    # Чередующаяся заливка: нечётные колонки (B=2, D=4...) с заливкой 5%
                    if col_idx % 2 == 0:  # B=2, D=4, F=6...
                        cell.fill = self._stripe_fill_5

                # Итого по сотруднику
                ws.cell(current_row, total_col).value = employee_total
                ws.cell(current_row, total_col).font = bold_font
                ws.cell(current_row, total_col).border = self._thin_border
                ws.cell(current_row, total_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Отклонение от индивидуальной нормы (итого - норма * ставка)
                deviation = employee_total - employee_norm
                ws.cell(current_row, deviation_col).value = deviation
                ws.cell(current_row, deviation_col).font = bold_font
                ws.cell(current_row, deviation_col).border = self._thin_border
                ws.cell(current_row, deviation_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                grand_total += employee_total
                current_row += 1

            # Последняя строка данных сотрудников (для толстых границ)
            last_data_row = current_row - 1
            totals_row = current_row  # Строка "Сумма по объекту"

            # === Строка сумм по проектам (с толстыми границами) ===
            # Колонка A - заливка 15%
            ws.cell(current_row, 1).value = "Сумма по объекту"
            ws.cell(current_row, 1).font = bold_font
            ws.cell(current_row, 1).fill = self._stripe_fill_15
            ws.cell(current_row, 1).border = self._thick_border
            ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Колонки шифров/категорий - заливка 15%
            for col_idx, column in enumerate(all_columns, start=2):
                cell = ws.cell(current_row, col_idx)
                total = column_totals.get(column, 0)
                if total > 0:
                    cell.value = total
                cell.font = bold_font
                cell.fill = self._stripe_fill_15
                cell.border = self._thick_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Общий итог (без заливки)
            ws.cell(current_row, total_col).value = grand_total
            ws.cell(current_row, total_col).font = bold_font
            ws.cell(current_row, total_col).border = self._thick_border
            ws.cell(current_row, total_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Общее отклонение (сумма часов - сумма индивидуальных норм)
            total_deviation = grand_total - total_expected_hours
            ws.cell(current_row, deviation_col).value = total_deviation
            ws.cell(current_row, deviation_col).font = bold_font
            ws.cell(current_row, deviation_col).border = self._thick_border
            ws.cell(current_row, deviation_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            current_row += 1

            # === Строка названий проектов ===
            ws.cell(current_row, 1).value = "Название объекта"
            ws.cell(current_row, 1).font = italic_font
            ws.cell(current_row, 1).border = self._thin_border
            ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for col_idx, column in enumerate(all_columns, start=2):
                cell = ws.cell(current_row, col_idx)
                # Для специальных категорий название пустое
                name = names.get(column, "")
                cell.value = name
                cell.font = default_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self._thin_border

            # Пустые ячейки в строке названий для колонок Итого и Отклонение
            ws.cell(current_row, total_col).border = self._thin_border
            ws.cell(current_row, deviation_col).border = self._thin_border

            # === Настройка ширины колонок ===
            # ВАЖНО: Используем прямое присвоение .width вместо ColumnDimension()
            # ColumnDimension с min/max параметрами предназначен для ЧТЕНИЯ из Excel,
            # а не для ЗАПИСИ. Простое присвоение .width - единственный надёжный способ.
            # См: https://joshuahunter.com/posts/setting-column-width-with-openpyxl/

            # Колонка A - ФИО
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['A'].customWidth = True

            # Шифры проектов - ширина 12, специальные категории - по словарю
            for col_idx, column in enumerate(all_columns, start=2):
                col_letter = get_column_letter(col_idx)
                # Проверяем, есть ли фиксированная ширина для этой категории
                if column in self.CATEGORY_COLUMN_WIDTHS:
                    width = self.CATEGORY_COLUMN_WIDTHS[column]
                else:
                    # Шифры проектов (и все остальное) - ширина 12
                    width = 12
                ws.column_dimensions[col_letter].width = width
                ws.column_dimensions[col_letter].customWidth = True

            # Итого и Отклонение - фиксированные ширины
            total_letter = get_column_letter(total_col)
            deviation_letter = get_column_letter(deviation_col)
            ws.column_dimensions[total_letter].width = self.TOTAL_COL_WIDTH
            ws.column_dimensions[total_letter].customWidth = True
            ws.column_dimensions[deviation_letter].width = self.DEVIATION_COL_WIDTH
            ws.column_dimensions[deviation_letter].customWidth = True

            # === Высота строки с названиями ===
            ws.row_dimensions[current_row].height = 60

            # === Объединение ячеек ===
            # A1:C1 - заголовок "План-график загрузки"
            ws.merge_cells('A1:C1')

            # A4:A5 - заголовок "ФИО сотрудника"
            ws.merge_cells('A4:A5')

            # B4:(последний шифр проекта) - заголовок "Шифр объекта"
            # Шифры проектов идут с колонки 2 (B) до колонки 1 + len(codes)
            if codes:
                last_code_col = 1 + len(codes)  # Последняя колонка с шифром проекта
                last_code_letter = get_column_letter(last_code_col)
                ws.merge_cells(f'B4:{last_code_letter}4')

            # Объединение строк 4-5 для специальных категорий
            # first_category_col уже определено выше
            for col_idx in range(first_category_col, len(all_columns) + 2):
                col_letter = get_column_letter(col_idx)
                ws.merge_cells(f'{col_letter}4:{col_letter}5')

            # Объединение строк 4-5 для "Итого часов за период"
            total_letter = get_column_letter(total_col)
            ws.merge_cells(f'{total_letter}4:{total_letter}5')

            # Объединение строк 4-5 для "Отклонение от нормы"
            deviation_letter = get_column_letter(deviation_col)
            ws.merge_cells(f'{deviation_letter}4:{deviation_letter}5')

            # === Толстые внешние границы для диапазонов ===
            # Последняя колонка с данными (спец.категории)
            last_cat_col = len(all_columns) + 1

            # Диапазон A4:последняя_категория - внешние толстые границы
            # (основная таблица данных до строки перед "Сумма по объекту")
            self._apply_outer_thick_border(ws, 4, 1, last_data_row, last_cat_col)

            # Колонка "Итого часов за период" - внешние толстые границы
            self._apply_outer_thick_border(ws, 4, total_col, last_data_row, total_col)

            # Колонка "Отклонение от нормы" - внешние толстые границы
            self._apply_outer_thick_border(ws, 4, deviation_col, last_data_row, deviation_col)

            # Закрепляем первую колонку (A - ФИО) и 5 строк заголовка
            ws.freeze_panes = 'B6'

            # Сохраняем файл с динамическим именем
            output_filename = f"Сводный табель Отдел архитектуры_и ПМ_{year}_{month:02d}.xlsx"
            output_path = Path(output_folder) / output_filename
            wb.save(str(output_path))
            wb.close()

            log_info(
                f"Fsm_6_1_5: Сводный табель создан: {output_path.name}, "
                f"сотрудников: {len(timesheets)}, шифров: {len(codes)}, категорий: {len(categories)}, "
                f"общий итог часов: {grand_total}"
            )

            return str(output_path)

        except Exception as e:
            log_error(f"Fsm_6_1_5: Ошибка генерации: {e}")
            return None
