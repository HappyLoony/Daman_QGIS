# -*- coding: utf-8 -*-
"""
Генератор объединенного табеля (Объединенный_табель.xlsx).

Собирает данные из всех валидных табелей сотрудников в один файл,
сохраняя структуру исходного шаблона.
"""

import warnings
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any, TYPE_CHECKING

from Daman_QGIS.utils import log_info, log_warning, log_error
from Daman_QGIS.managers.reference.submodules import ProductionCalendarManager
from .Fsm_6_1_3_parser import (
    TimesheetData, ProjectRow, SpecialCategoryRow,
    safe_read_excel, COL_CODE, COL_NAME, COL_TOTAL,
    COL_DAYS_START, COL_DAYS_END
)

# Ленивый импорт openpyxl для избежания ошибки np.float при загрузке плагина
if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Шрифт по умолчанию для генерируемых файлов
DEFAULT_FONT_NAME = "Times New Roman"
DEFAULT_FONT_SIZE = 11


class MergedTimesheetGenerator:
    """Генератор объединенного табеля."""

    # Имя выходного файла
    OUTPUT_FILENAME = "Объединенный_табель.xlsx"

    # Строки заголовка: пропускаем первые 2 строки (пометки для сотрудников)
    HEADER_START_ROW = 3  # Начинаем копирование с 3-й строки
    HEADER_END_ROW = 9    # Заканчиваем на 9-й строке

    def __init__(self):
        self._template_wb: Optional['Workbook'] = None
        self._template_filepath: Optional[str] = None
        self._styles_initialized = False

        # Стили (инициализируются лениво)
        self._stripe_fill_5: Optional['PatternFill'] = None
        self._stripe_fill_15: Optional['PatternFill'] = None
        self._thin_side: Optional['Side'] = None
        self._thick_side: Optional['Side'] = None
        self._thin_border: Optional['Border'] = None
        self._thick_border: Optional['Border'] = None
        self._center_alignment: Optional['Alignment'] = None

        # Кэш выходных дней для заливки (заполняется в generate)
        self._holidays: set = set()  # {день_месяца, ...}

    def _init_styles(self) -> None:
        """Инициализировать стили openpyxl (ленивая загрузка)."""
        if self._styles_initialized:
            return

        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        # Стили заливок
        # "Белый, фон 1, более темный оттенок 5%" = D9D9D9 (светло-серый)
        # "Белый, фон 1, более темный оттенок 15%" = BFBFBF (серый)
        self._stripe_fill_5 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        self._stripe_fill_15 = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

        # Границы
        self._thin_side = Side(style='thin')
        self._thick_side = Side(style='medium')

        self._thin_border = Border(
            left=self._thin_side,
            right=self._thin_side,
            top=self._thin_side,
            bottom=self._thin_side
        )

        self._thick_border = Border(
            left=self._thick_side,
            right=self._thick_side,
            top=self._thick_side,
            bottom=self._thick_side
        )

        # Стандартное выравнивание
        self._center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        self._styles_initialized = True

    def _apply_outer_thick_border(self, ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
        """
        Применить толстые внешние границы к диапазону ячеек.

        Args:
            ws: Worksheet
            start_row: Начальная строка
            start_col: Начальная колонка
            end_row: Конечная строка
            end_col: Конечная колонка
        """
        from openpyxl.styles import Border
        self._init_styles()

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row, col)

                # Определяем какие стороны должны быть толстыми
                left = self._thick_side if col == start_col else self._thin_side
                right = self._thick_side if col == end_col else self._thin_side
                top = self._thick_side if row == start_row else self._thin_side
                bottom = self._thick_side if row == end_row else self._thin_side

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def _load_template(self, template_filepath: str) -> bool:
        """
        Загрузить шаблон из одного из табелей.

        Args:
            template_filepath: Путь к файлу-шаблону

        Returns:
            True если успешно
        """
        # Патч numpy для совместимости со старыми версиями openpyxl
        from .Fsm_6_1_3_parser import _patch_numpy_float
        _patch_numpy_float()

        from openpyxl import load_workbook

        try:
            # Подавляем предупреждение openpyxl о Data Validation
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", message="Data Validation extension")
                # Открываем БЕЗ read_only чтобы иметь доступ к column_dimensions
                self._template_wb = load_workbook(template_filepath, data_only=True)
            self._template_filepath = template_filepath
            log_info(f"Fsm_6_1_4: Шаблон загружен из {Path(template_filepath).name}")
            return True
        except Exception as e:
            log_error(f"Fsm_6_1_4: Не удалось загрузить шаблон: {template_filepath}: {e}")
            return False

    # Колонки для пропуска при копировании заголовка
    # E(5)=Номер договора, F(6)=Дата договора, G(7)=Заказчик
    # Эти колонки не нужны в объединенном табеле
    SKIP_HEADER_COLS = {5, 6, 7}

    def _copy_header(self, source_ws, target_ws) -> None:
        """
        Копировать заголовок из шаблона.

        Пропускает первые 2 строки (пометки для сотрудников).
        Очищает колонки E, F, G (Номер договора, Дата договора, Заказчик).

        Args:
            source_ws: Исходный лист
            target_ws: Целевой лист
        """
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        # Копируем строки 3-9 из шаблона в строки 1-7 результата
        target_row = 1
        for source_row in range(self.HEADER_START_ROW, self.HEADER_END_ROW + 1):
            for col in range(1, COL_DAYS_END + 1):
                source_cell = source_ws.cell(source_row, col)
                target_cell = target_ws.cell(target_row, col)

                # Пропускаем значения колонок E, F, G (договор, дата, заказчик)
                if col in self.SKIP_HEADER_COLS:
                    target_cell.value = None
                elif isinstance(source_cell.value, datetime):
                    target_cell.value = source_cell.value.strftime("%d.%m.%Y")
                else:
                    target_cell.value = source_cell.value

                # Копируем форматирование если есть
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic
                    )

                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical,
                        wrap_text=source_cell.alignment.wrap_text
                    )

            target_row += 1

        # Ширины колонок задаются явно в generate() ПОСЛЕ delete_cols,
        # т.к. delete_cols не сдвигает column_dimensions (баг openpyxl),
        # а чтение из шаблона не работает для range-based dimensions

    def _write_project_row(
        self,
        ws,
        row: int,
        project: ProjectRow,
        days_in_month: int,
        end_day: Optional[int] = None
    ) -> None:
        """
        Записать строку проекта.

        Args:
            ws: Worksheet
            row: Номер строки
            project: Данные проекта
            days_in_month: Количество дней в месяце
            end_day: Последний день для записи (включительно). Если None - все дни.
        """
        from openpyxl.styles import Font
        self._init_styles()

        default_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)

        # Вид работ (колонка B)
        cell_work_type = ws.cell(row, 2)
        if project.work_type:
            cell_work_type.value = project.work_type
        cell_work_type.font = default_font
        cell_work_type.alignment = self._center_alignment
        cell_work_type.border = self._thin_border

        # Шифр проекта
        cell_code = ws.cell(row, COL_CODE)
        cell_code.value = project.code
        cell_code.font = default_font
        cell_code.alignment = self._center_alignment
        cell_code.border = self._thin_border

        # Название объекта
        cell_name = ws.cell(row, COL_NAME)
        if project.name:
            cell_name.value = project.name
        cell_name.font = default_font
        cell_name.alignment = self._center_alignment
        cell_name.border = self._thin_border

        # Итого часов (только до end_day)
        total_hours = 0.0
        limit_day = end_day if end_day is not None else days_in_month
        for day, hours in project.daily_hours.items():
            if day <= limit_day:
                total_hours += hours

        cell_total = ws.cell(row, COL_TOTAL)
        cell_total.value = total_hours
        cell_total.font = default_font
        cell_total.alignment = self._center_alignment
        cell_total.border = self._thin_border

        # Часы по дням (только до end_day)
        for day in range(1, days_in_month + 1):
            col = COL_DAYS_START + day - 1
            cell = ws.cell(row, col)
            if day <= limit_day:
                hours = project.daily_hours.get(day)
                if hours:
                    cell.value = hours
            cell.font = default_font
            cell.alignment = self._center_alignment
            cell.border = self._thin_border
            # Заливка для выходных и праздничных дней
            if day in self._holidays:
                cell.fill = self._stripe_fill_5

    def _write_special_category_row(
        self,
        ws,
        row: int,
        category: SpecialCategoryRow,
        days_in_month: int,
        end_day: Optional[int] = None
    ) -> None:
        """
        Записать строку специальной категории.

        Args:
            ws: Worksheet
            row: Номер строки
            category: Данные категории
            days_in_month: Количество дней в месяце
            end_day: Последний день для записи (включительно). Если None - все дни.
        """
        from openpyxl.styles import Font
        self._init_styles()

        default_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)

        # Колонка B - пустая для спец.категорий
        cell_b = ws.cell(row, 2)
        cell_b.border = self._thin_border

        # Название категории (объединяем ячейки C:D)
        cell_cat = ws.cell(row, COL_CODE)
        cell_cat.value = category.category
        cell_cat.font = default_font
        cell_cat.alignment = self._center_alignment
        cell_cat.border = self._thin_border

        # Ячейка COL_NAME - будет объединена с COL_CODE
        cell_name = ws.cell(row, COL_NAME)
        cell_name.border = self._thin_border

        # Объединяем C:D для специальной категории
        ws.merge_cells(
            start_row=row, start_column=COL_CODE,
            end_row=row, end_column=COL_NAME
        )

        # Итого часов (только до end_day)
        total_hours = 0.0
        limit_day = end_day if end_day is not None else days_in_month
        for day, hours in category.daily_hours.items():
            if day <= limit_day:
                total_hours += hours

        cell_total = ws.cell(row, COL_TOTAL)
        cell_total.value = total_hours
        cell_total.font = default_font
        cell_total.alignment = self._center_alignment
        cell_total.border = self._thin_border

        # Часы по дням (только до end_day)
        for day in range(1, days_in_month + 1):
            col = COL_DAYS_START + day - 1
            cell = ws.cell(row, col)
            if day <= limit_day:
                hours = category.daily_hours.get(day)
                if hours:
                    cell.value = hours
            cell.font = default_font
            cell.alignment = self._center_alignment
            cell.border = self._thin_border
            # Заливка для выходных и праздничных дней
            if day in self._holidays:
                cell.fill = self._stripe_fill_5

    def _write_employee_total(
        self,
        ws,
        row: int,
        total_hours: float,
        days_in_month: int
    ) -> None:
        """
        Записать итоговую строку сотрудника.

        Args:
            ws: Worksheet
            row: Номер строки
            total_hours: Итого часов
            days_in_month: Количество дней в месяце
        """
        from openpyxl.styles import Font
        self._init_styles()

        bold_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)

        # Колонка B - пустая для строки итого
        cell_b = ws.cell(row, 2)
        cell_b.border = self._thin_border
        cell_b.fill = self._stripe_fill_15

        # Итого
        cell_code = ws.cell(row, COL_CODE)
        cell_code.value = "Итого"
        cell_code.font = bold_font
        cell_code.alignment = self._center_alignment
        cell_code.border = self._thin_border
        cell_code.fill = self._stripe_fill_15

        # Пустая ячейка COL_NAME
        cell_name = ws.cell(row, COL_NAME)
        cell_name.border = self._thin_border
        cell_name.fill = self._stripe_fill_15

        # Итого часов
        cell_total = ws.cell(row, COL_TOTAL)
        cell_total.value = total_hours
        cell_total.font = bold_font
        cell_total.alignment = self._center_alignment
        cell_total.border = self._thin_border
        cell_total.fill = self._stripe_fill_15

        # Заливка ячеек дней в строке итого
        for day in range(1, days_in_month + 1):
            col = COL_DAYS_START + day - 1
            cell = ws.cell(row, col)
            cell.border = self._thin_border
            cell.fill = self._stripe_fill_15

    def generate(
        self,
        timesheets: List[TimesheetData],
        output_folder: str,
        end_day: Optional[int] = None
    ) -> Optional[str]:
        """
        Сгенерировать объединенный табель.

        Args:
            timesheets: Список данных табелей
            output_folder: Папка для сохранения
            end_day: Последний день для расчётов (включительно).
                     Если None - используются все дни месяца.

        Returns:
            Путь к созданному файлу или None при ошибке
        """
        # Патч numpy для совместимости со старыми версиями openpyxl
        from .Fsm_6_1_3_parser import _patch_numpy_float
        _patch_numpy_float()

        from openpyxl import Workbook
        from openpyxl.styles import Font

        if not timesheets:
            log_warning("Fsm_6_1_4: Нет табелей для объединения")
            return None

        log_info(f"Fsm_6_1_4: Генерация объединенного табеля для {len(timesheets)} сотрудников")

        # Инициализируем стили
        self._init_styles()

        # Сортируем по фамилии
        timesheets_sorted = sorted(timesheets, key=lambda t: t.surname.lower())

        # Загружаем шаблон из первого табеля
        if not self._load_template(timesheets_sorted[0].filepath):
            return None

        try:
            # Создаем новую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "табель"

            # Получаем лист шаблона
            template_sheets = self._template_wb.sheetnames
            template_ws = self._template_wb[template_sheets[0]]

            # Копируем заголовок
            self._copy_header(template_ws, ws)

            # Определяем параметры
            first_timesheet = timesheets_sorted[0]
            year = first_timesheet.year
            month = first_timesheet.month

            import calendar
            from datetime import date
            days_in_month = calendar.monthrange(year, month)[1]

            # Кэшируем выходные и праздничные дни для заливки
            calendar_manager = ProductionCalendarManager()
            self._holidays = set()
            for day in range(1, days_in_month + 1):
                if calendar_manager.is_holiday(date(year, month, day)):
                    self._holidays.add(day)

            # Начинаем с 8-й строки (заголовок занимает 7 строк после пропуска 2-х служебных)
            current_row = 8

            # Записываем данные каждого сотрудника
            grand_total_hours = 0.0
            bold_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)

            for ts in timesheets_sorted:
                employee_start_row = current_row

                # Первая строка сотрудника - ФИО в колонке A, первый проект в остальных
                cell_fio = ws.cell(current_row, 1)
                cell_fio.value = ts.fio
                cell_fio.font = bold_font
                cell_fio.alignment = self._center_alignment
                cell_fio.border = self._thin_border

                # Проекты (колонка B заполняется в _write_project_row)
                for project in ts.projects:
                    self._write_project_row(ws, current_row, project, days_in_month, end_day)
                    current_row += 1

                # Специальные категории
                for category in ts.special_categories:
                    self._write_special_category_row(ws, current_row, category, days_in_month, end_day)
                    current_row += 1

                # Итого по сотруднику (с учётом end_day)
                employee_total = ts.get_hours_until_day(end_day) if end_day else ts.total_hours
                self._write_employee_total(ws, current_row, employee_total, days_in_month)
                grand_total_hours += employee_total
                employee_end_row = current_row

                # Объединяем ячейки ФИО по вертикали (только колонка A)
                if employee_end_row > employee_start_row:
                    ws.merge_cells(
                        start_row=employee_start_row,
                        start_column=1,
                        end_row=employee_end_row,
                        end_column=1
                    )

                current_row += 1

                # Пустая строка между сотрудниками
                current_row += 1

            # Общий итог
            cell_total_label = ws.cell(current_row, COL_CODE)
            cell_total_label.value = "ОБЩИЙ ИТОГ"
            cell_total_label.font = bold_font
            cell_total_label.alignment = self._center_alignment
            cell_total_label.border = self._thick_border
            cell_total_label.fill = self._stripe_fill_15

            cell_total_b = ws.cell(current_row, 2)
            cell_total_b.border = self._thick_border
            cell_total_b.fill = self._stripe_fill_15

            cell_total_name = ws.cell(current_row, COL_NAME)
            cell_total_name.border = self._thick_border
            cell_total_name.fill = self._stripe_fill_15

            cell_total_value = ws.cell(current_row, COL_TOTAL)
            cell_total_value.value = grand_total_hours
            cell_total_value.font = bold_font
            cell_total_value.alignment = self._center_alignment
            cell_total_value.border = self._thick_border
            cell_total_value.fill = self._stripe_fill_15

            # Заливка ячеек дней в строке общего итога
            for day in range(1, days_in_month + 1):
                col = COL_DAYS_START + day - 1
                cell = ws.cell(current_row, col)
                cell.border = self._thick_border
                cell.fill = self._stripe_fill_15

            # Удаляем пустые колонки E, F, G (договор, дата, заказчик) со сдвигом влево
            ws.delete_cols(5, 3)

            # Ширины колонок задаём ПОСЛЕ delete_cols, т.к.:
            # 1) delete_cols сдвигает ячейки, но НЕ column_dimensions
            # 2) Чтение из шаблона ненадёжно (range-based dimensions)
            # После удаления E/F/G структура: A=ФИО, B=Вид работ, C=Шифр,
            # D=Название, E=Итого, F+=Дни (1..days_in_month)
            from openpyxl.utils import get_column_letter

            ws.column_dimensions['A'].width = 35   # ФИО
            ws.column_dimensions['B'].width = 15   # Вид работ
            ws.column_dimensions['C'].width = 12   # Шифр
            ws.column_dimensions['D'].width = 20   # Название
            ws.column_dimensions['E'].width = 10   # Итого
            for day_col in range(6, 6 + days_in_month):
                ws.column_dimensions[get_column_letter(day_col)].width = 4.5

            # Закрепляем первые 2 колонки (A, B) и 7 строк заголовка
            ws.freeze_panes = 'C8'

            # Сохраняем файл
            output_path = Path(output_folder) / self.OUTPUT_FILENAME
            wb.save(str(output_path))
            wb.close()

            log_info(
                f"Fsm_6_1_4: Объединенный табель создан: {output_path.name}, "
                f"сотрудников: {len(timesheets)}, часов: {grand_total_hours}"
            )

            return str(output_path)

        except Exception as e:
            log_error(f"Fsm_6_1_4: Ошибка генерации: {e}")
            return None

        finally:
            if self._template_wb:
                self._template_wb.close()
                self._template_wb = None
