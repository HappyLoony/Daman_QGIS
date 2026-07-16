# -*- coding: utf-8 -*-
"""
Субмодуль Fsm_4_2_T_6_1_norm_v2 - Тесты F_6_1 (по-дневная норма с cap,
ревизия 9 + ревизия 10: отгул выкупает норму, не нейтрализует).

Машинная планка реализации плана
`documentation/plans/_PLAN_F_6_1_part_rate_vacation_2026-06-22.md`.

Ревизия 10 (два множества, R10.2): NEUTRALIZING={Отпуск,Больничный}
нейтрализуют норму подённо с cap; BUYOUT={Отгул} выкупает норму (ст. 128) --
в факте (employee_total), НЕ в норме -> полностью уменьшает норму. Различие
выкуп<->нейтрализация ВИДНО только когда cap кусает (отгул-часы > day_norm):
D1 (rate 0.5, отгул@8 > day_norm=4 -> старый 0 / новый +4),
D3 (rate 1.0 предпраздничный, отгул@8 > day_norm=7 -> старый 0 / новый +1).
rate 1.0 обычный день (отгул@8 = day_norm=8, cap НЕ кусает) тождественен в
обоих режимах -> для инъекции дефекта БЕСПОЛЕЗЕН (test_08 старой ревизии был
именно таким -> слеп). Эталоны инъекции ОБЯЗАНЫ быть cap-кусающими.

КЛЮЧЕВОЕ ОТЛИЧИЕ ОТ Fsm_4_2_T_6_1.py (§7, FIX-8):
- employee_total берётся из РЕАЛЬНЫХ daily_hours через
  SummaryTimesheetGenerator._build_data_matrix (НЕ хардкод) -> deviation
  считается полной связкой _build_data_matrix -> compute_employee_norm ->
  round(total-norm). Прежние тесты маскировали Дефект C хардкодом employee_total;
- кейсы невидимых символов / колонки размещения (§7.12) прогоняются через
  РЕАЛЬНЫЙ parse_timesheet (временный xlsx), а не прямой SpecialCategoryRow --
  иначе тест обошёл бы парсер-гейт классификации (C0/C0b);
- инвариант факт<->норма на неполном периоде (§7.13): absence-доля employee_total
  == absence_total == Σ absence_by_day.

Норма тестируется ОФЛАЙН через детерминированный фейк ProductionCalendarManager
(_FakeCalendar) с управляемыми is_workday/is_shortened_day -- без сети.

Эталон календаря по умолчанию (_make_calendar_20wd): 20 рабочих дней (1..20),
без предпраздничных. year=2026, month=1 (31 день; дни 21..31 не рабочие).
  day_norm(rate=1.0) = 8, day_norm(rate=0.5) = 4.
Формула (§4, по-дневная с cap):
  day_norm(d)   = max(8*rate - (1 если предпраздничный иначе 0), 0)   # 0 если не рабочий
  excused(d)    = min(absence_on_d, day_norm(d))
  work_norm     = Σ по рабочим d (day_norm - excused)
  employee_norm = work_norm + Σ absence_by_day
  deviation     = round(employee_total - employee_norm, 2)
"""

import os
import tempfile
from datetime import date, timedelta


class _FakeCalendar:
    """Детерминированный фейк ProductionCalendarManager для офлайн-тестов нормы.

    Не ходит в сеть. Рабочие/предпраздничные дни задаются множествами номеров.
    Совместим по интерфейсу с compute_employee_norm И с generate()
    (get_work_hours_for_period / get_work_hours_for_month для шапки нормы).

    Args:
        workdays: множество номеров рабочих дней месяца (1..31).
        shortened: множество номеров предпраздничных (сокращённых) дней.
        raise_runtime: если True -- is_workday/is_shortened_day СРАЗУ бросают
                       RuntimeError (эмуляция полностью недоступного календаря,
                       F-07 graceful None).
        fail_on_scan_start: если задано (N>=1) -- is_workday бросает RuntimeError
                       на N-м вызове со check_date.day == 1. Каждый полный обход
                       периода (шапка get_work_hours_for_period/_for_month и
                       каждый compute_employee_norm) начинается с дня 1, поэтому
                       счётчик day==1 = номер начавшегося скана. Позволяет
                       календарю УСПЕШНО отработать шапку + первых K сотрудников
                       и упасть на (K+1)-м -> тест дыры §4d (mid-run all-or-nothing,
                       FIX-9): первые K строк уже получили число до флага.
    """

    def __init__(self, workdays, shortened=None, raise_runtime=False,
                 fail_on_scan_start=None):
        self._workdays = set(workdays)
        self._shortened = set(shortened or [])
        self._raise_runtime = raise_runtime
        self._fail_on_scan_start = fail_on_scan_start
        self._scan_start_count = 0

    def is_workday(self, check_date: date) -> bool:
        if self._raise_runtime:
            raise RuntimeError("Производственный календарь недоступен (фейк)")
        if self._fail_on_scan_start is not None and check_date.day == 1:
            self._scan_start_count += 1
            if self._scan_start_count >= self._fail_on_scan_start:
                raise RuntimeError(
                    "Производственный календарь стал недоступен в процессе (фейк)"
                )
        return check_date.day in self._workdays

    def is_shortened_day(self, check_date: date) -> bool:
        if self._raise_runtime:
            raise RuntimeError("Производственный календарь недоступен (фейк)")
        return check_date.day in self._shortened

    def is_holiday(self, check_date: date) -> bool:
        """Выходной/праздник = не рабочий день (реплика Msm_4_22.is_holiday).

        Использует is_workday -> уважает raise_runtime/fail_on_scan_start
        (validate_overtime/validate_absence_placement ловят RuntimeError, F-07).
        """
        return not self.is_workday(check_date)

    # --- Интерфейс, требуемый generate() для расчёта шапки "Норма часов" ---

    def get_work_hours_for_period(self, start_date: date, end_date: date,
                                  hours_per_day: int = 8) -> float:
        """Норма часов за период (реплика Msm_4_22 через is_workday/is_shortened).

        Обход день-за-днём с дня start_date -- согласовано с fail_on_scan_start:
        первый вызов приходится на день 1 периода (шапка стартует с month_start).
        """
        total = 0.0
        current = start_date
        while current <= end_date:
            if self.is_workday(current):
                if self.is_shortened_day(current):
                    total += hours_per_day - 1
                else:
                    total += hours_per_day
            current += timedelta(days=1)
        return total

    def get_work_hours_for_month(self, year: int, month: int,
                                 hours_per_day: int = 8) -> float:
        """Норма часов за весь месяц (реплика Msm_4_22)."""
        import calendar as _cal
        num_days = _cal.monthrange(year, month)[1]
        return self.get_work_hours_for_period(
            date(year, month, 1), date(year, month, num_days), hours_per_day
        )


def _make_calendar_20wd(shortened=None, raise_runtime=False):
    """Фейк-календарь: 20 рабочих дней (1..20), без предпраздничных по умолчанию."""
    return _FakeCalendar(
        workdays=set(range(1, 21)),
        shortened=shortened,
        raise_runtime=raise_runtime
    )


def _make_timesheet(fio="Иванов Иван Иванович", projects=None,
                    special_categories=None, total_hours=0.0, year=2026, month=1):
    """Создать TimesheetData программно (для формульных кейсов, не парсер-гейта)."""
    from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_3_parser import TimesheetData

    return TimesheetData(
        filepath="test.xlsx",
        filename="Иванов_01.xlsx",
        fio=fio,
        month_start=date(year, month, 1),
        month_end=date(year, month, 28),
        year=year,
        month=month,
        projects=projects or [],
        special_categories=special_categories or [],
        total_hours=total_hours
    )


def _make_category(name, daily_hours, row_number=20):
    """SpecialCategoryRow с подённой детализацией (total = Σ daily)."""
    from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_3_parser import SpecialCategoryRow

    return SpecialCategoryRow(
        row_number=row_number,
        category=name,
        total_hours=float(sum(daily_hours.values())),
        daily_hours=dict(daily_hours)
    )


def _make_project(code, daily_hours, row_number=10):
    """ProjectRow с подённой детализацией (total = Σ daily)."""
    from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_3_parser import ProjectRow

    return ProjectRow(
        row_number=row_number,
        code=code,
        name=f"Проект {code}",
        total_hours=float(sum(daily_hours.values())),
        daily_hours=dict(daily_hours)
    )


# Колонки шаблона табеля (совпадают с Fsm_6_1_3_parser)
_COL_CODE = 3      # C -- шифр проекта
_COL_NAME = 4      # D -- название / категория (альтернативная колонка)
_COL_TOTAL = 8     # H -- итого
_COL_DAYS_START = 9  # I -- день 1
_ROW_DATA_START = 10  # первая строка данных (= A10 ФИО)


def _write_timesheet_xlsx(path, fio, rows, year=2026, month=1,
                          end_day=28, put_itogo=True):
    """Записать реальный xlsx-табель по шаблону парсера (для §7.12 парсер-гейта).

    rows: список dict:
        {"col": "C"|"D", "value": <шифр или категория>,
         "name": <опц. значение колонки D когда col=="C">,
         "daily": {день: часы}}
    B4 -- дата начала месяца (для year/month), A10 -- ФИО.
    Строка "Итого" (H) добавляется после данных, чтобы find_data_bounds
    корректно нашёл границу.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "табель"

    # B4 -- начало месяца (datetime -> парсер извлекает year/month)
    ws.cell(4, 2).value = date(year, month, 1)
    ws.cell(4, 3).value = date(year, month, end_day)
    # A10 -- ФИО
    ws.cell(_ROW_DATA_START, 1).value = fio

    r = _ROW_DATA_START
    for row in rows:
        col = _COL_CODE if row["col"] == "C" else _COL_NAME
        ws.cell(r, col).value = row["value"]
        # Опциональное имя в D, если категория/шифр в C
        if row["col"] == "C" and row.get("name") is not None:
            ws.cell(r, _COL_NAME).value = row["name"]
        daily = row.get("daily", {})
        total = 0.0
        for day, hours in daily.items():
            ws.cell(r, _COL_DAYS_START + day - 1).value = hours
            total += hours
        ws.cell(r, _COL_TOTAL).value = total
        r += 1

    if put_itogo:
        ws.cell(r, 7).value = "Итого"  # колонка G (в диапазоне 5..8)

    wb.save(path)
    wb.close()


class TestF61NormV2:
    """Тесты F_6_1: по-дневная норма с cap + полная связка факт<->норма."""

    def __init__(self, iface, logger):
        self.iface = iface
        self.logger = logger
        self._tempdir = None

    def run_all_tests(self):
        self.logger.section("ТЕСТ F_6_1 v2: по-дневная норма с cap")
        try:
            self._tempdir = tempfile.mkdtemp(prefix="daman_t61_")

            # Формульные кейсы (полная связка через _build_data_matrix)
            self.test_01_rate05_vacation_neutralized()
            self.test_02_rate1_sick_plus_work_overtime()
            self.test_03_rate1_partial_absence_day()
            self.test_04_rate05_sick_plus_work_overtime()
            self.test_05_rate05_full_vacation()
            self.test_06_shortened_day_full_hour()
            self.test_07_absence_on_shortened_day()
            self.test_08_otgul_buyout_differentiating_D2()
            self.test_09_kp_as_work_reduces_deficit()
            self.test_10_year_boundary_not_rejected()
            self.test_11_month_rejection()
            self.test_11a_blast_radius_gate()
            self.test_12_invisible_chars_and_column_c()
            self.test_13_fact_norm_invariant_partial_period()
            self.test_14_calendar_unavailable()
            self.test_14b_midrun_all_or_nothing()
            self.test_14c_single_employee_missing_ts()
            self.test_15_total_daily_desync_rejected()
            self.test_16_total_deviation_sum_of_rounded()
            self.test_17_manager_norm_meaningful()
            self.test_18_malformed_inputs_no_crash()

            # Ревизия 10: выкуп отгула + валидация размещения
            self.test_19_otgul_buyout_injection_D1()
            self.test_20_otgul_buyout_injection_D3_shortened()
            self.test_21_otgul_plus_vacation_two_classes_O4()
            self.test_22_overtime_absence_categories_skipped()
            self.test_23_otgul_parser_gate_invisible_chars()
            self.test_24_absence_placement_rejects_anomalies()
            self.test_25_absence_placement_valid_controls()
            self.test_26_absence_placement_calendar_runtime_failsoft()
            self.test_27_overtime_absence_no_phantom_valid_otgul()
            self.test_28_validator_summary_consistency_otgul()

        except Exception as e:
            self.logger.error(f"Критическая ошибка: {e}")
            import traceback
            self.logger.data("Traceback", traceback.format_exc())
        finally:
            self._cleanup_tempdir()

        self.logger.summary()

    def _cleanup_tempdir(self):
        if not self._tempdir:
            return
        try:
            import shutil
            shutil.rmtree(self._tempdir, ignore_errors=True)
        except Exception:
            pass

    # ------------------------------------------------------------------
    # Полная связка: _build_data_matrix -> compute_employee_norm -> deviation
    # ------------------------------------------------------------------

    def _full_chain(self, ts, rate, calendar, end_day=None, year=2026, month=1,
                    calendar_unavailable=False):
        """Полная связка расчёта. Возвращает (employee_total, employee_norm, deviation).

        employee_total -- из РЕАЛЬНОГО _build_data_matrix (НЕ хардкод).
        employee_norm -- compute_employee_norm через absence_by_day из тех же daily.
        """
        from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_5_summary import (
            SummaryTimesheetGenerator
        )
        from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_6_norm import (
            compute_employee_norm, absence_by_day
        )

        gen = SummaryTimesheetGenerator()
        codes, categories = gen._collect_unique_codes([ts])
        employees_data, _ = gen._build_data_matrix(
            [ts], codes, categories, end_day=end_day
        )
        # employee_total как в generate(): Σ значений hours_by_column
        _, hours_by_column = employees_data[0]
        employee_total = sum(hours_by_column.values())

        absence_map = absence_by_day(ts, end_day)
        employee_norm = compute_employee_norm(
            ts, absence_map, rate, end_day, calendar, year, month,
            calendar_unavailable
        )
        if employee_norm is None:
            return employee_total, None, None
        deviation = round(employee_total - employee_norm, 2)
        return employee_total, employee_norm, deviation

    # ------------------------------------------------------------------
    # Реальный generate() -> запись xlsx -> чтение ячеек "Отклонение".
    # Отличает FIX-7 (итог = Σ round(per-employee)) и FIX-9 (mid-run
    # all-or-nothing) от изолированных вызовов helper: тесты дёргают
    # производственный путь построения файла и проверяют РЕАЛЬНЫЕ ячейки.
    # ------------------------------------------------------------------

    def _run_generate(self, timesheets, calendar, report_end_date,
                      employee_rates, calendar_unavailable=False,
                      subdir="gen"):
        """Запустить РЕАЛЬНЫЙ SummaryTimesheetGenerator.generate() в темп-папку.

        Возвращает (output_path, returned_calendar_unavailable).
        """
        from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_5_summary import (
            SummaryTimesheetGenerator
        )
        out_folder = os.path.join(self._tempdir, subdir)
        os.makedirs(out_folder, exist_ok=True)
        gen = SummaryTimesheetGenerator()
        path, cal_unavail = gen.generate(
            timesheets, out_folder,
            report_end_date=report_end_date,
            employee_rates=employee_rates,
            calendar_manager=calendar,
            calendar_unavailable=calendar_unavailable
        )
        return path, cal_unavail

    def _read_deviation_column(self, xlsx_path):
        """Прочитать колонку 'Отклонение' построенного сводного табеля.

        Находит колонку по заголовку 'Отклонение от нормы' (строка 4), строку
        итога -- по 'Сумма по объекту' в колонке A. Возвращает dict:
            {"per_employee": [(fio, dev_value), ...],
             "total": <значение итоговой ячейки 'Отклонение'>}
        dev_value/total = None если ячейка пуста (не рассчитано).
        """
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, data_only=True)
        try:
            ws = wb.active
            # Колонка "Отклонение от нормы" (заголовок в строке 4)
            deviation_col = None
            max_col = ws.max_column
            for col in range(1, max_col + 1):
                val = ws.cell(4, col).value
                if val and "Отклонение" in str(val):
                    deviation_col = col
                    break
            if deviation_col is None:
                raise AssertionError("Колонка 'Отклонение от нормы' не найдена")

            # Строка итога "Сумма по объекту" в колонке A
            totals_row = None
            for row in range(6, ws.max_row + 1):
                a_val = ws.cell(row, 1).value
                if a_val and "Сумма по объекту" in str(a_val):
                    totals_row = row
                    break
            if totals_row is None:
                raise AssertionError("Строка 'Сумма по объекту' не найдена")

            per_employee = []
            for row in range(6, totals_row):
                fio = ws.cell(row, 1).value
                dev = ws.cell(row, deviation_col).value
                per_employee.append((fio, dev))

            total = ws.cell(totals_row, deviation_col).value
            return {"per_employee": per_employee, "total": total}
        finally:
            wb.close()

    def test_01_rate05_vacation_neutralized(self):
        """§7.1 (кейс A): 0.5, 5дн отпуск@8 + 15 раб@4 -> deviation 0 (разоблачает Дефект C)."""
        self.logger.section("1. rate=0.5 отпуск нейтрализован (кейс A)")
        try:
            vac_daily = {d: 8.0 for d in range(1, 6)}          # дни 1-5, отпуск
            work_daily = {d: 4.0 for d in range(6, 21)}        # дни 6-20, работа
            ts = _make_timesheet(
                projects=[_make_project("25-П-1", work_daily)],
                special_categories=[_make_category("Отпуск", vac_daily)]
            )
            total, norm, dev = self._full_chain(ts, 0.5, _make_calendar_20wd())
            self.logger.check(
                total == 100.0 and norm == 100.0 and dev == 0.0,
                "0.5: total=100, norm=100, dev=0 (глоб.формула дала бы +20)",
                f"total={total}, norm={norm}, dev={dev} (ожидалось 100/100/0)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_02_rate1_sick_plus_work_overtime(self):
        """§7.2 (кейс B): 1.0, 8ч больн.+2ч работа + 19 норм@8 -> +2 (Q2)."""
        self.logger.section("2. rate=1.0 больничный+работа -> переработка (кейс B)")
        try:
            sick_daily = {1: 8.0}
            work_daily = {1: 2.0}
            work_daily.update({d: 8.0 for d in range(2, 21)})  # дни 2-20 по 8
            ts = _make_timesheet(
                projects=[_make_project("25-П-1", work_daily)],
                special_categories=[_make_category("Больничный", sick_daily)]
            )
            total, norm, dev = self._full_chain(ts, 1.0, _make_calendar_20wd())
            self.logger.check(
                total == 162.0 and norm == 160.0 and dev == 2.0,
                "1.0: total=162, norm=160, dev=+2 (работа в день отсутствия)",
                f"total={total}, norm={norm}, dev={dev} (ожидалось 162/160/+2)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_03_rate1_partial_absence_day(self):
        """§7.3 (кейс C): 1.0, 4ч отпуск+4ч работа + 19@8 -> 0 (cap; без cap +4)."""
        self.logger.section("3. rate=1.0 частичный день отсутствия (кейс C)")
        try:
            vac_daily = {1: 4.0}
            work_daily = {1: 4.0}
            work_daily.update({d: 8.0 for d in range(2, 21)})
            ts = _make_timesheet(
                projects=[_make_project("25-П-1", work_daily)],
                special_categories=[_make_category("Отпуск", vac_daily)]
            )
            total, norm, dev = self._full_chain(ts, 1.0, _make_calendar_20wd())
            self.logger.check(
                total == 160.0 and norm == 160.0 and dev == 0.0,
                "1.0: 4ч отпуск+4ч работа -> dev=0 (cap min(4,8)=4)",
                f"total={total}, norm={norm}, dev={dev} (ожидалось 160/160/0)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_04_rate05_sick_plus_work_overtime(self):
        """§7.4 (кейс D): 0.5, 8ч больн.+1ч работа + 19 раб@4 -> +1 (Q2 на дробной)."""
        self.logger.section("4. rate=0.5 больничный+работа -> переработка (кейс D)")
        try:
            sick_daily = {1: 8.0}
            work_daily = {1: 1.0}
            work_daily.update({d: 4.0 for d in range(2, 21)})  # 19 дней по 4 = 76
            ts = _make_timesheet(
                projects=[_make_project("25-П-1", work_daily)],
                special_categories=[_make_category("Больничный", sick_daily)]
            )
            total, norm, dev = self._full_chain(ts, 0.5, _make_calendar_20wd())
            # work_norm: день1 excused=min(8,4)=4 -> 0; дни2-20 = 76. absence_total=8.
            # norm=84. total = (8+1) + 76 = 85. dev=+1.
            self.logger.check(
                total == 85.0 and norm == 84.0 and dev == 1.0,
                "0.5: 8ч больн.+1ч работа -> dev=+1",
                f"total={total}, norm={norm}, dev={dev} (ожидалось 85/84/+1)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_05_rate05_full_vacation(self):
        """§7.5 (кейс E): 0.5, весь период отпуск@8 -> 0."""
        self.logger.section("5. rate=0.5 полный отпуск (кейс E)")
        try:
            vac_daily = {d: 8.0 for d in range(1, 21)}  # 20 дней по 8 = 160
            ts = _make_timesheet(
                special_categories=[_make_category("Отпуск", vac_daily)]
            )
            total, norm, dev = self._full_chain(ts, 0.5, _make_calendar_20wd())
            # work_norm: каждый день excused=min(8,4)=4 -> 0. absence_total=160.
            # norm=160. total=160. dev=0.
            self.logger.check(
                total == 160.0 and norm == 160.0 and dev == 0.0,
                "0.5: полный отпуск -> dev=0 (norm=absence_total=160)",
                f"total={total}, norm={norm}, dev={dev} (ожидалось 160/160/0)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_06_shortened_day_full_hour(self):
        """§7.6 (кейс F): предпраздничный -> минус ПОЛНЫЙ 1 час (ст.95), rate 0.5 и 1.0."""
        self.logger.section("6. Предпраздничный день -1ч полный (кейс F)")
        try:
            # Без отсутствий, день 20 -- предпраздничный
            cal = _make_calendar_20wd(shortened={20})
            ts1 = _make_timesheet()
            _, norm1, _ = self._full_chain(ts1, 1.0, cal)
            self.logger.check(
                norm1 == 159.0,
                "rate=1.0, предпраздничный -> 159 (19*8 + 7)",
                f"norm={norm1} (ожидалось 159.0)"
            )
            ts05 = _make_timesheet()
            _, norm05, _ = self._full_chain(ts05, 0.5, cal)
            self.logger.check(
                norm05 == 79.0,
                "rate=0.5, предпраздничный -> 79 (19*4 + 3, минус полный 1ч)",
                f"norm={norm05} (ожидалось 79.0)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_07_absence_on_shortened_day(self):
        """§7.7: absence x предпраздничный (кейс G) + частичный предпраздничный (H, H2)."""
        self.logger.section("7. Отсутствие/работа на предпраздничном (G/H/H2)")
        try:
            cal = _make_calendar_20wd(shortened={20})

            # Кейс G: 1.0, 8ч отпуск на предпраздничном дне 20 + 19@8
            vac_g = {20: 8.0}
            work_g = {d: 8.0 for d in range(1, 20)}
            ts_g = _make_timesheet(
                projects=[_make_project("25-П-1", work_g)],
                special_categories=[_make_category("Отпуск", vac_g)]
            )
            tg, ng, dg = self._full_chain(ts_g, 1.0, cal)
            self.logger.check(
                tg == 160.0 and ng == 160.0 and dg == 0.0,
                "G: excused=min(8,7)=7, день обнулён, dev=0",
                f"total={tg}, norm={ng}, dev={dg} (ожидалось 160/160/0)"
            )

            # Кейс H: 1.0, предпраздничный 4ч отпуск+3ч работа + 19@8
            vac_h = {20: 4.0}
            work_h = {20: 3.0}
            work_h.update({d: 8.0 for d in range(1, 20)})
            ts_h = _make_timesheet(
                projects=[_make_project("25-П-1", work_h)],
                special_categories=[_make_category("Отпуск", vac_h)]
            )
            th, nh, dh = self._full_chain(ts_h, 1.0, cal)
            self.logger.check(
                th == 159.0 and nh == 159.0 and dh == 0.0,
                "H: day_norm(пр)=7, excused=min(4,7)=4, work=3, dev=0",
                f"total={th}, norm={nh}, dev={dh} (ожидалось 159/159/0)"
            )

            # Кейс H2: то же, 4ч отпуск + 5ч работа -> +2
            vac_h2 = {20: 4.0}
            work_h2 = {20: 5.0}
            work_h2.update({d: 8.0 for d in range(1, 20)})
            ts_h2 = _make_timesheet(
                projects=[_make_project("25-П-1", work_h2)],
                special_categories=[_make_category("Отпуск", vac_h2)]
            )
            th2, nh2, dh2 = self._full_chain(ts_h2, 1.0, cal)
            self.logger.check(
                th2 == 161.0 and nh2 == 159.0 and dh2 == 2.0,
                "H2: переработка на частичном предпраздничном, dev=+2",
                f"total={th2}, norm={nh2}, dev={dh2} (ожидалось 161/159/+2)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_08_otgul_buyout_differentiating_D2(self):
        """§7.8 (ревизия 10, D2): отгул ВЫКУПАЕТ норму (cap-кусающий эталон).

        РАЗЛИЧАЮЩИЙ эталон D2 (rate 0.5, отгул@8 + переработка): старый режим
        (отгул нейтрализуется) дал бы +8, новый (отгул выкупает) даёт +12. rate
        0.5 обязателен: cap min(отгул8, day_norm4) кусает -> различие видно.

        ИНЪЕКЦИЯ ДЕФЕКТА (канон-21): на ТЕКУЩЕМ (нейтрализация) коде этот эталон
        FAIL (dev=+8, ассерт ждёт +12); после ревизии 10 PASS (dev=+12). См.
        implementation-notes run-dir. D2 -- ОБЫЧНЫЙ overtime-регресс (та же ось
        rate<1.0 что D1, знак "+"), не единственный механизм инъекции (D1+D3 --
        два разных механизма кусания cap).
        """
        self.logger.section("8. Отгул выкупает норму, D2 (rate 0.5, cap кусает)")
        try:
            # rate 0.5, 20 раб.дней -> норма (новая) = 80 (нет absence).
            # отгул@8 день1; работа: день1=4, день2=8, дни3-20=4 (work_fact=84).
            # total = 84 (работа) + 8 (отгул) = 92.
            # НОВЫЙ (выкуп): norm=80 -> dev = 92-80 = +12.
            # СТАРЫЙ (нейтрализ.): отгул в absence, absence_total=8;
            #   work_norm: день1 excused=min(8,4)=4 -> 0; дни2-20=76; work_norm=76;
            #   norm_old = 76+8 = 84 -> dev = 92-84 = +8. Различает (+8 vs +12).
            otgul = {1: 8.0}
            work = {1: 4.0, 2: 8.0}
            work.update({d: 4.0 for d in range(3, 21)})  # дни3-20 = 18*4 = 72
            ts = _make_timesheet(
                projects=[_make_project("25-П-1", work)],
                special_categories=[_make_category("Отгул", otgul)]
            )
            total, norm, dev = self._full_chain(ts, 0.5, _make_calendar_20wd())
            self.logger.check(
                total == 92.0 and norm == 80.0 and dev == 12.0,
                "D2: отгул выкупил норму -> total=92, norm=80, dev=+12 "
                "(старый режим дал бы +8)",
                f"total={total}, norm={norm}, dev={dev} (ожидалось 92/80/+12)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_09_kp_as_work_reduces_deficit(self):
        """§7.9: КП как работа -- входит в факт, уменьшает |deviation| при недоборе, НЕ нейтрализуется."""
        self.logger.section("9. КП как работа (не отсутствие)")
        try:
            # 1.0: работа 15 дней@8 = 120 + КП 5 дней@8 = 40. Итог 160.
            # КП НЕ нейтрализуется, идёт как шифр в факт. norm=160 (все 20 раб.дней).
            # dev = 160 - 160 = 0. КП закрыл дефицит как РАБОТА (не как absence).
            work = {d: 8.0 for d in range(1, 16)}
            kp = {d: 8.0 for d in range(16, 21)}
            ts = _make_timesheet(
                projects=[_make_project("25-П-1", work)],
                special_categories=[_make_category("Коммерческие предложения", kp)]
            )
            total, norm, dev = self._full_chain(ts, 1.0, _make_calendar_20wd())
            self.logger.check(
                total == 160.0 and norm == 160.0 and dev == 0.0,
                "КП как работа: total=160 (вкл. КП), norm=160, dev=0",
                f"total={total}, norm={norm}, dev={dev} (ожидалось 160/160/0)"
            )
            # Контроль: если бы КП нейтрализовался как absence, norm стала бы
            # > 160 (добавили бы КП обратно). Проверяем что absence_by_day КП НЕ ловит.
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_6_norm import absence_by_day
            amap = absence_by_day(ts, None)
            self.logger.check(
                sum(amap.values()) == 0.0,
                "КП НЕ в absence_by_day (Σ absence = 0)",
                f"Σ absence = {sum(amap.values())} (ожидалось 0.0 -- КП не absence)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_10_year_boundary_not_rejected(self):
        """§7.10: B4-год == GUI-год декабря -> НЕ отбракован."""
        self.logger.section("10. Граница года (декабрь) не отбракован")
        try:
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_2_validator import (
                TimesheetValidator
            )
            # Табель за декабрь 2025, GUI-год 2025, месяц 12 -> совпадает.
            ts = _make_timesheet(year=2025, month=12)
            validator = TimesheetValidator(target_month=12, target_year=2025)
            result = validator.validate_month(ts)
            self.logger.check(
                result.is_valid,
                "B4=12.2025, target=12.2025 -> validate_month валиден",
                f"is_valid={result.is_valid} (ожидалось True)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_11_month_rejection(self):
        """§7.11: отбраковка месяца (B4!=GUI год+месяц; B4=None; битый B4 в том же году)."""
        self.logger.section("11. Отбраковка месяца (add_error)")
        try:
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_2_validator import (
                TimesheetValidator
            )
            validator = TimesheetValidator(target_month=1, target_year=2026)

            # (а) месяц не совпадает: B4=02.2026, target=01.2026
            ts_wrong = _make_timesheet(year=2026, month=2)
            r1 = validator.validate_month(ts_wrong)
            self.logger.check(
                (not r1.is_valid) and r1.errors_count >= 1,
                "B4=02.2026 != target 01.2026 -> add_error (отбраковка)",
                f"is_valid={r1.is_valid}, errors={r1.errors_count} (ожидалось False/>=1)"
            )

            # (б) B4 не распарсился (month_start=None) -> add_error
            ts_none = _make_timesheet(year=2026, month=1)
            ts_none.month_start = None
            r2 = validator.validate_month(ts_none)
            self.logger.check(
                (not r2.is_valid) and r2.errors_count >= 1,
                "month_start=None -> add_error (отбраковка, FIX-5)",
                f"is_valid={r2.is_valid}, errors={r2.errors_count} (ожидалось False/>=1)"
            )

            # (в) битый B4 в ТОМ ЖЕ году: парсер дал бы year=now(). Симулируем
            # month_start=None (что и есть маркер битого B4) при year==target ->
            # всё равно отбракован через None-check-первым (R5-5/OPT-011).
            ts_broken = _make_timesheet(year=2026, month=1)
            ts_broken.month_start = None  # битый B4 -> None-check ловит первым
            r3 = validator.validate_month(ts_broken)
            self.logger.check(
                not r3.is_valid,
                "битый B4 (None) в том же году -> отбракован (None-check первым)",
                f"is_valid={r3.is_valid} (ожидалось False)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_11a_blast_radius_gate(self):
        """§7.11a: невалид по НЕ-месячной причине -> workdays/overtime warnings ПОДАВЛЕНЫ."""
        self.logger.section("11a. Blast-radius гейта FIX-6 (OPT-019)")
        try:
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_2_validator import (
                TimesheetValidator
            )
            # Табель валиден по имени/ФИО/месяцу (01.2026 = target), но невалиден
            # по рассинхрону total/daily: total=80, daily пуст -> consistency-ошибка.
            # Гейт `if not is_valid: return` ДО validate_workdays -> НЕТ warnings
            # "рабочий день не заполнен" (иначе на GUI-календаре их было бы много).
            #
            # Праймим кэши валидатора -> filename/fio проходят ОФЛАЙН (без сети
            # к Base_employee.json), чтобы изолировать именно гейт FIX-6.
            ts = _make_timesheet(year=2026, month=1, total_hours=80.0,
                                 special_categories=[_make_category("Отпуск", {})])
            ts.total_hours = 80.0  # daily пуст -> рассинхрон

            validator = TimesheetValidator(target_month=1, target_year=2026,
                                           end_day=20)
            fake_emp = {
                "last_name": "Иванов", "first_name": "Иван",
                "middle_name": "Иванович", "rate": 1.0
            }
            validator._employees_cache = [fake_emp]
            validator._surnames_cache = {"иванов"}
            # valid_project_codes не нужен (у ts нет проектов), но праймим на всякий
            validator._valid_project_codes = set()

            result = validator.validate(ts)
            self.logger.check(
                not result.is_valid,
                "рассинхрон total/daily -> табель невалиден",
                f"is_valid={result.is_valid} (ожидалось False)"
            )
            # Гейт стоит ПОСЛЕ consistency, поэтому ошибка рассинхрона present
            has_consistency_err = any(
                "не сходится" in m.message for m in result.messages
            )
            self.logger.check(
                has_consistency_err,
                "ошибка рассинхрона total/daily достигнута (гейт после неё)",
                f"сообщения: {[m.message for m in result.messages]}"
            )
            # Ни одного warning про незаполненный рабочий день (подавлены гейтом)
            has_workday_warn = any(
                "рабочий день" in m.message.lower() for m in result.messages
            )
            self.logger.check(
                not has_workday_warn,
                "workdays/overtime warnings ПОДАВЛЕНЫ (гейт до validate_workdays)",
                f"обнаружены вторичные warnings: {[m.message for m in result.messages]}"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_12_invisible_chars_and_column_c(self):
        """§7.12: невидимые символы + регистр + колонка размещения ЧЕРЕЗ РЕАЛЬНЫЙ парсер."""
        self.logger.section("12. Невидимые символы / колонка C (реальный парсер, C0/C0b)")
        try:
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_3_parser import (
                parse_timesheet
            )
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_6_norm import absence_by_day

            zwsp = "​"       # zero-width space
            nbsp_narrow = " "  # narrow no-break space

            # (а) "Отпуск<zwsp>" в колонке D -> должен классифицироваться как
            #     отсутствие (не проект-шифр), попасть в факт И норму.
            # (б) "Коммерческие<narrow-nbsp>предложения" (внутренний nbsp) в D ->
            #     КП в факт (иначе КП-часы теряются).
            # (в) "коммерческие предложения" (lowercase) -> КП (case-fold).
            # (г) absence-строка в колонке C (OPT-016) -> НЕ dropped.
            path = os.path.join(self._tempdir, "Тестов_01.xlsx")
            _write_timesheet_xlsx(
                path, "Тестов Тест Тестович",
                rows=[
                    {"col": "D", "value": f"Отпуск{zwsp}", "daily": {1: 8.0, 2: 8.0}},
                    {"col": "D", "value": f"Коммерческие{nbsp_narrow}предложения",
                     "daily": {3: 8.0}},
                    {"col": "D", "value": "коммерческие предложения", "daily": {4: 8.0}},
                    {"col": "C", "value": f"Больничный{zwsp}", "daily": {5: 8.0}},
                ],
                year=2026, month=1, end_day=28
            )
            ts = parse_timesheet(path)
            self.logger.check(
                ts is not None,
                "parse_timesheet вернул табель",
                "parse_timesheet вернул None"
            )
            if ts is None:
                return

            # Собираем категории по нормализованному имени (АККУМУЛЯЦИЯ:
            # несколько строк с одной нормализованной категорией суммируются,
            # как в _build_data_matrix -- КП с внутр.nbsp и lowercase = одна).
            from Daman_QGIS.utils import normalize_for_classification
            cats = {}
            for c in ts.special_categories:
                key = normalize_for_classification(c.category).lower()
                cats[key] = cats.get(key, 0.0) + sum(c.daily_hours.values())
            # (а) Отпуск с zwsp -> распознан как категория (не проект)
            self.logger.check(
                cats.get("отпуск", 0.0) == 16.0,
                "Отпуск<zwsp> -> категория отпуск 16ч (не проект-шифр)",
                f"cats={cats} (ожидалось отпуск=16)"
            )
            # (а+) НЕ попал в projects как шифр
            self.logger.check(
                len(ts.projects) == 0,
                "Отпуск<zwsp> НЕ в projects (C0 нормализация парсер-гейта)",
                f"projects={[p.code for p in ts.projects]} (ожидалось пусто)"
            )
            # (б)+(в) КП (внутренний nbsp + lowercase) суммарно 16ч
            self.logger.check(
                cats.get("коммерческие предложения", 0.0) == 16.0,
                "КП (внутр.nbsp + lowercase) -> 16ч в факте (C0/C3 case-fold)",
                f"cats={cats} (ожидалось КП=16)"
            )
            # (г) Больничный из колонки C -> НЕ dropped
            self.logger.check(
                cats.get("больничный", 0.0) == 8.0,
                "Больничный в колонке C -> НЕ dropped (C0b)",
                f"cats={cats} (ожидалось больничный=8)"
            )
            # absence_by_day видит отпуск(16) + больничный(8) = 24 (КП НЕ absence)
            amap = absence_by_day(ts, None)
            self.logger.check(
                sum(amap.values()) == 24.0,
                "absence_by_day = 24 (отпуск16 + больничный8, КП не absence)",
                f"Σ absence = {sum(amap.values())} (ожидалось 24.0)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_13_fact_norm_invariant_partial_period(self):
        """§7.13 (ревизия 10 / A1): инвариант факт<->норма через NEUTRALIZING.

        Инвариант `Σ(NEUTRALIZING факта) == absence_total == Σ absence_by_day` на
        НЕПОЛНОМ периоде. КРИТИЧНО (rev-code-r6 ISSUE-001): отбор идёт по
        NEUTRALIZING (Отпуск/Больничный), НЕ ABSENCE -- absence_by_day в ревизии 10
        исключает отгул, значит и fact_absence должен считаться по NEUTRALIZING,
        иначе инвариант ложно упадёт. Отгул в employee_total, НЕ в absence_by_day.
        """
        self.logger.section("13. Инвариант факт<->норма (NEUTRALIZING, неполный период)")
        try:
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_5_summary import (
                SummaryTimesheetGenerator
            )
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_6_norm import (
                absence_by_day, compute_employee_norm, NEUTRALIZING
            )
            from Daman_QGIS.utils import normalize_for_classification

            # Табель: отпуск дни 1-10 по 8ч (NEUTRALIZING) + отгул день 3 8ч
            # (BUYOUT -- в факте, НЕ в absence_by_day). end_day=5 (середина).
            # Инвариант по NEUTRALIZING: fact_neutral(до5) == absence_total(до5)
            #   == Σ absence_by_day (все = отпуск дни1-5 = 40; отгул НЕ входит).
            vac_daily = {d: 8.0 for d in range(1, 11)}
            otgul_daily = {3: 8.0}  # отгул в факте, но НЕ в absence_by_day
            ts = _make_timesheet(
                special_categories=[
                    _make_category("Отпуск", vac_daily),
                    _make_category("Отгул", otgul_daily),
                ]
            )
            end_day = 5

            gen = SummaryTimesheetGenerator()
            codes, categories = gen._collect_unique_codes([ts])
            employees_data, _ = gen._build_data_matrix(
                [ts], codes, categories, end_day=end_day
            )
            _, hours_by_column = employees_data[0]
            employee_total = sum(hours_by_column.values())

            amap = absence_by_day(ts, end_day)
            absence_total = sum(amap.values())

            # neutralizing-доля employee_total: сумма ТОЛЬКО NEUTRALIZING-колонок
            # факта (A1: НЕ ABSENCE -- отгул исключён из инварианта нормы).
            neutralizing_set = {
                normalize_for_classification(c).lower() for c in NEUTRALIZING
            }
            fact_neutral = sum(
                h for col, h in hours_by_column.items()
                if normalize_for_classification(col).lower() in neutralizing_set
            )
            # отгул-доля факта (для контроля: он ЕСТЬ в факте, но НЕ в норме)
            fact_otgul = sum(
                h for col, h in hours_by_column.items()
                if normalize_for_classification(col).lower() == "отгул"
            )

            norm = compute_employee_norm(
                ts, amap, 0.5, end_day, _make_calendar_20wd(), 2026, 1, False
            )
            deviation = round(employee_total - norm, 2)

            # neutralizing-инвариант: fact_neutral == absence_total == Σamap == 40
            # (отпуск дни1-5 = 40; отгул НЕ входит).
            self.logger.check(
                fact_neutral == absence_total == sum(amap.values()) == 40.0,
                "инвариант (NEUTRALIZING): fact_neutral == absence_total == "
                "Σ absence_by_day == 40 (отгул НЕ в absence_by_day)",
                f"fact_neutral={fact_neutral}, absence_total={absence_total} "
                f"(ожидалось 40 все)"
            )
            # отгул ЕСТЬ в факте (день3 <= 5): fact_otgul=8. Целевая асимметрия.
            self.logger.check(
                fact_otgul == 8.0,
                "отгул в employee_total (день3 <= end_day) -- выкуп через факт",
                f"fact_otgul={fact_otgul} (ожидалось 8.0)"
            )
            # Отклонение: norm(до5,0.5) = work_norm(5дн, все excused отпуском) +
            #   absence_total(40) = 0 + 40 = 40. Отгул@8 день3 в факте -> выкуп
            #   на 8ч: dev = total - norm. total(до5) = отпуск40 + отгул8 = 48
            #   (отпуск дни1-5=40, отгул день3=8). norm=40. dev = 48-40 = +8.
            self.logger.check(
                deviation == 8.0,
                "отгул выкупил норму на неполном периоде -> dev=+8 "
                "(старый режим нейтрализовал бы -> dev=0)",
                f"total={employee_total}, norm={norm}, dev={deviation} (ожидалось +8)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_14_calendar_unavailable(self):
        """§7.14: недоступность календаря (флаг / RuntimeError -> None), helper-уровень."""
        self.logger.section("14. Недоступность календаря -> None (helper)")
        try:
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_6_norm import (
                compute_employee_norm, absence_by_day
            )
            ts = _make_timesheet()
            amap = absence_by_day(ts, None)

            # (а) явный флаг -> None
            r_flag = compute_employee_norm(
                ts, amap, 1.0, None, _make_calendar_20wd(), 2026, 1, True
            )
            self.logger.check(
                r_flag is None,
                "calendar_unavailable=True -> None",
                f"результат={r_flag} (ожидался None)"
            )
            # (б) RuntimeError из календаря в процессе -> None (graceful)
            r_rt = compute_employee_norm(
                ts, amap, 1.0, None, _make_calendar_20wd(raise_runtime=True),
                2026, 1, False
            )
            self.logger.check(
                r_rt is None,
                "RuntimeError -> None (graceful, F-07)",
                f"результат={r_rt} (ожидался None)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_14b_midrun_all_or_nothing(self):
        """§7.14b (FIX-9/§4d): mid-run сбой календаря -> 'Отклонение' пусто у ВСЕХ.

        Прежний test дёргал compute_employee_norm в изоляции -> не проверял дыру:
        первые K сотрудников УЖЕ получили числовое отклонение в ячейку ДО того,
        как календарь упал и был выставлен флаг. Здесь -- РЕАЛЬНЫЙ generate() с
        _FakeCalendar, который успешно отвечает для шапки + первых 2 сотрудников
        и бросает RuntimeError на 3-м (fail_on_scan_start=5: скан-старты
        #1,#2 = шапка get_work_hours_for_period/_for_month, #3,#4 = сотр.1,2,
        #5 = сотр.3 -> падение). Проверяем по построенному xlsx: (а) ВСЕ
        per-employee ячейки 'Отклонение' пусты, включая первых 2 (пост-цикловая
        очистка deviation_cells), не только хвост; (б) итог пуст; (в) generate()
        вернул calendar_unavailable=True.
        """
        self.logger.section("14b. Mid-run all-or-nothing (FIX-9, реальный xlsx)")
        try:
            work = {d: 8.0 for d in range(1, 21)}  # 160ч, норма 160 -> dev=0

            def _emp(fio):
                return _make_timesheet(
                    fio=fio, projects=[_make_project("25-П-1", dict(work))]
                )

            # 3 сотрудника; фамилии сортируются -> порядок обхода:
            # Алов(1) -> Бунин(2) -> Волков(3). Падение на 3-м.
            employees = [_emp("Алов А.А."), _emp("Бунин Б.Б."), _emp("Волков В.В.")]

            cal = _FakeCalendar(workdays=set(range(1, 21)), fail_on_scan_start=5)
            report_end = date(2026, 1, 31)

            path, cal_unavail = self._run_generate(
                employees, cal, report_end, employee_rates={},
                subdir="fix9_midrun"
            )
            # (в) флаг недоступности возвращён
            self.logger.check(
                cal_unavail is True,
                "generate() вернул calendar_unavailable=True (mid-run падение)",
                f"cal_unavail={cal_unavail} (ожидалось True)"
            )
            self.logger.check(
                path is not None,
                "файл всё равно построен (F-07: пустые ячейки, не крах)",
                f"path={path} (ожидался не-None)"
            )
            if not path:
                return

            data = self._read_deviation_column(path)
            per_emp = data["per_employee"]
            # (а) ВСЕ per-employee 'Отклонение' пусты (вкл. первых 2)
            self.logger.check(
                len(per_emp) == 3 and all(v is None for _, v in per_emp),
                "ВСЕ 3 ячейки 'Отклонение' пусты (пост-цикловая очистка первых K)",
                f"per_employee={per_emp} (ожидались все None)"
            )
            # (б) итог пуст
            self.logger.check(
                data["total"] is None,
                "итог 'Отклонение' пуст (any_norm_unavailable/флаг)",
                f"total={data['total']} (ожидался None)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_14c_single_employee_missing_ts(self):
        """§7.14c: сотрудник без ts (нет в timesheets_by_fio) -> пусто ТОЛЬКО ему.

        Дыра иного рода, чем 14b: не сбой календаря, а отсутствие TimesheetData
        для одного ФИО (ts_for_norm is None -> employee_norm None -> его ячейка
        пуста, ветка generate() line 531-533), остальные считаются штатно. Итог
        пуст (any_norm_unavailable), но per-employee ячейки ОСТАЛЬНЫХ содержат
        число, и пост-цикловой очистки НЕТ (calendar_unavailable=False).

        Механизм: monkeypatch _build_data_matrix добавляет 'призрачную' строку
        с ФИО, которого нет в timesheets_by_fio -> generate() на этой строке
        получает ts_for_norm=None. Это точечно бьёт нужную ветку без сети/парсера.
        """
        self.logger.section("14c. Один сотрудник без ts -> пусто только ему")
        try:
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_5_summary import (
                SummaryTimesheetGenerator
            )

            work = {d: 8.0 for d in range(1, 21)}  # 160ч, норма 160 -> dev=0
            ts_ok = _make_timesheet(
                fio="Гусев Г.Г.", projects=[_make_project("25-П-1", dict(work))]
            )
            ghost_fio = "Призрак П.П."  # ФИО без соответствующего TimesheetData

            gen = SummaryTimesheetGenerator()
            orig_build = gen._build_data_matrix

            def _patched_build(timesheets, codes, categories, end_day=None):
                emp_data, totals = orig_build(
                    timesheets, codes, categories, end_day=end_day
                )
                # 'Призрачная' строка: ФИО отсутствует в timesheets_by_fio,
                # значит ts_for_norm станет None (целевая ветка).
                ghost_hours = dict(emp_data[0][1]) if emp_data else {}
                emp_data.append((ghost_fio, ghost_hours))
                return emp_data, totals

            gen._build_data_matrix = _patched_build  # type: ignore[method-assign]

            out_folder = os.path.join(self._tempdir, "fix9_missing")
            os.makedirs(out_folder, exist_ok=True)
            path, cal_unavail = gen.generate(
                [ts_ok], out_folder,
                report_end_date=date(2026, 1, 31),
                employee_rates={},
                calendar_manager=_make_calendar_20wd(),
                calendar_unavailable=False
            )
            self.logger.check(
                path is not None and not cal_unavail,
                "generate() построил файл, календарь доступен (не mid-run сбой)",
                f"path={path}, cal_unavail={cal_unavail}"
            )
            if not path:
                return

            data = self._read_deviation_column(path)
            per_emp = data["per_employee"]
            by_fio = {fio: v for fio, v in per_emp}
            # Валидный сотрудник -> число (dev=0), призрак -> пусто
            self.logger.check(
                by_fio.get("Гусев Г.Г.") is not None
                and round(by_fio.get("Гусев Г.Г."), 2) == 0.0,
                "валидный сотрудник -> 'Отклонение' = 0 (посчитан)",
                f"per_employee={per_emp} (ожидалось Гусев=0)"
            )
            self.logger.check(
                ghost_fio in by_fio and by_fio.get(ghost_fio) is None,
                "сотрудник без ts -> 'Отклонение' пусто ТОЛЬКО ему",
                f"per_employee={per_emp} (ожидалось {ghost_fio}=None)"
            )
            # Итог пуст: any_norm_unavailable=True (одна норма None)
            self.logger.check(
                data["total"] is None,
                "итог пуст (any_norm_unavailable из-за одной None-нормы)",
                f"total={data['total']} (ожидался None)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_15_total_daily_desync_rejected(self):
        """§7.15: рассинхрон total/daily -> отвергнут validate_total_consistency."""
        self.logger.section("15. Рассинхрон total/daily -> ошибка")
        try:
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_2_validator import (
                TimesheetValidator
            )
            # total=80, daily пуст -> расхождение 80 > 0.5 -> ошибка.
            ts = _make_timesheet(total_hours=80.0,
                                 special_categories=[_make_category("Отпуск", {})])
            ts.total_hours = 80.0
            validator = TimesheetValidator(target_month=1, target_year=2026)
            consistency = validator.validate_total_consistency(ts)
            self.logger.check(
                (not consistency.is_valid) and consistency.errors_count >= 1,
                "total=80 без daily -> validate_total_consistency ошибка",
                f"is_valid={consistency.is_valid}, errors={consistency.errors_count}"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_16_total_deviation_sum_of_rounded(self):
        """§7.16: инвариант итога (FIX-7) -- РЕАЛЬНЫЙ generate() + чтение xlsx.

        Прежний тест был тавтологией на литералах (round(0.33)+round(0.34)==0.67):
        прошёл бы БЕЗ фикса. Здесь два сотрудника с per-employee отклонением +0.014
        каждый: помесячное округление ячейки даёт round(0.014,2)=0.01, а итог по
        FIX-7 = Σ округлённых = 0.02. Наивная формула round(grand_total - Σнорм)
        = round(0.028,2) = 0.03. Тест зовёт generate(), читает ячейку итога
        'Отклонение' из построенного файла и требует 0.02 (== Σ округлённых
        per-employee ячеек) И != 0.03 (наивный round(Σ)). Тест, проходящий на
        старой формуле итога, невалиден.
        """
        self.logger.section("16. Итог отклонения = Σ round(per-employee) (FIX-7, реальный xlsx)")
        try:
            # Норма rate=1.0 = 160 (20 раб.дней). Факт 160.014 у каждого:
            # день1 = 8.014, дни2-20 = 8.0. deviation(round) = 0.01 у каждого.
            def _emp(fio):
                work = {1: 8.014}
                work.update({d: 8.0 for d in range(2, 21)})
                return _make_timesheet(
                    fio=fio, projects=[_make_project("25-П-1", work)]
                )

            ts1 = _emp("Иванов Иван Иванович")
            ts2 = _emp("Петров Пётр Петрович")

            report_end = date(2026, 1, 31)  # весь январь -> gui 2026/01, 20 раб.дней
            path, cal_unavail = self._run_generate(
                [ts1, ts2], _make_calendar_20wd(), report_end,
                employee_rates={}, subdir="fix7"
            )
            self.logger.check(
                path is not None and not cal_unavail,
                "generate() создал файл, календарь доступен",
                f"path={path}, cal_unavail={cal_unavail}"
            )
            if not path:
                return

            data = self._read_deviation_column(path)
            per_emp = data["per_employee"]
            total_cell = data["total"]

            # (1) каждая per-employee ячейка = round(0.014,2) = 0.01
            per_vals = [round(v, 2) if v is not None else None for _, v in per_emp]
            self.logger.check(
                len(per_vals) == 2 and all(v == 0.01 for v in per_vals),
                "per-employee 'Отклонение' = 0.01 у каждого (round помесячно)",
                f"per_employee={per_emp} (ожидалось [0.01, 0.01])"
            )

            # (2) итог == Σ округлённых per-employee ячеек (FIX-7)
            sum_of_rounded = round(sum(v for v in per_vals if v is not None), 2)
            self.logger.check(
                total_cell is not None and round(total_cell, 2) == sum_of_rounded == 0.02,
                "итог 'Отклонение' == Σ round(per-employee) == 0.02 (FIX-7)",
                f"total={total_cell}, Σround={sum_of_rounded} (ожидалось 0.02)"
            )

            # (3) итог НЕ равен наивному round(grand_total - Σнорм) = round(0.028) = 0.03
            #     -> тест реально отличает FIX-7 от старой формулы итога.
            naive = round((160.014 - 160.0) + (160.014 - 160.0), 2)  # = 0.03
            self.logger.check(
                total_cell is not None and round(total_cell, 2) != naive,
                "итог != наивный round(Σ сырых отклонений)=0.03 (различает FIX-7)",
                f"total={total_cell}, наивный={naive} (итог не должен быть 0.03)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_17_manager_norm_meaningful(self):
        """§7.17: руководитель участвует, норма осмысленна (тот же helper)."""
        self.logger.section("17. Руководитель -- норма осмысленна")
        try:
            work = {d: 8.0 for d in range(1, 21)}
            ts = _make_timesheet(fio="Божук Б.Б.",
                                 projects=[_make_project("25-П-1", work)])
            ts.is_manager = True
            total, norm, dev = self._full_chain(ts, 1.0, _make_calendar_20wd())
            self.logger.check(
                total == 160.0 and norm == 160.0 and dev == 0.0,
                "руководитель: total=160, norm=160, dev=0 (тот же расчёт)",
                f"total={total}, norm={norm}, dev={dev} (ожидалось 160/160/0)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_18_malformed_inputs_no_crash(self):
        """§7.18: невалид (битый B4, отрицательные/нечисловые часы, неизвестный шифр) -- без краха."""
        self.logger.section("18. Malformed-вход -- диагностика без краха")
        try:
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_3_parser import (
                parse_timesheet
            )
            # xlsx с текстом в B4 (битая дата) и нечисловыми часами -> парсер
            # не должен упасть, вернёт табель (year/month=now при битом B4).
            path = os.path.join(self._tempdir, "Кривов_01.xlsx")
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "табель"
            ws.cell(4, 2).value = "не дата"       # B4 битый
            ws.cell(10, 1).value = "Кривов К.К."   # A10 ФИО
            ws.cell(10, 3).value = "25-П-1"         # C шифр
            ws.cell(10, 9).value = "abc"            # I день1 нечисловой
            ws.cell(10, 10).value = -5              # J день2 отрицательный
            ws.cell(11, 7).value = "Итого"
            wb.save(path)
            wb.close()

            ts = parse_timesheet(path)
            self.logger.check(
                ts is not None,
                "parse_timesheet на битом входе -> табель (без краха)",
                "parse_timesheet упал (None)"
            )
            if ts is not None:
                # Нечисловые/отрицательные часы отфильтрованы (_parse_daily_hours
                # берёт только hours > 0)
                proj_days = ts.projects[0].daily_hours if ts.projects else {}
                self.logger.check(
                    all(h > 0 for h in proj_days.values()),
                    "нечисловые/отрицательные часы отфильтрованы (только >0)",
                    f"daily={proj_days} (ожидались только положительные)"
                )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    # ==================================================================
    # РЕВИЗИЯ 10: выкуп отгула, overtime absence-отбор, валидация размещения
    # ==================================================================

    def _make_validator(self, calendar, rate=1.0, fio="Иванов Иван Иванович",
                        target_month=1, target_year=2026, end_day=20):
        """TimesheetValidator с ВПРЫСНУТЫМ фейк-календарём и праймленным сотрудником.

        Изолирует validate_overtime/validate_absence_placement от сети: календарь
        -- _FakeCalendar, employee-кэш праймлен (ставка из dict, без Base_employee).
        """
        from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_2_validator import (
            TimesheetValidator
        )
        validator = TimesheetValidator(
            target_month=target_month, target_year=target_year, end_day=end_day
        )
        validator._calendar_manager = calendar  # впрыск фейка (обходит property)
        # Праймим сотрудника с нужной ставкой (fio -> rate) без сети.
        parts = fio.split()
        emp = {
            "last_name": parts[0] if parts else "Иванов",
            "first_name": parts[1] if len(parts) > 1 else "Иван",
            "middle_name": parts[2] if len(parts) > 2 else "Иванович",
            "rate": rate,
        }
        validator._employees_cache = [emp]
        validator._surnames_cache = {emp["last_name"].lower()}
        validator._valid_project_codes = set()
        return validator

    def test_19_otgul_buyout_injection_D1(self):
        """§R5/D1 (ИНЪЕКЦИЯ, механизм rate<1.0): 0.5, отгул@8 + работа76 -> 0/+4.

        ИНЪЕКЦИЯ ДЕФЕКТА (канон-21): на ТЕКУЩЕМ (нейтрализация) коде dev=0,
        ассерт ждёт +4 -> FAIL; после ревизии 10 -> PASS. Cap min(отгул8,4)
        кусает -> различает выкуп от нейтрализации.
        """
        self.logger.section("19. D1 инъекция: отгул выкуп (0.5, cap min(8,4) кусает)")
        try:
            # отгул@8 день1, работа дни2-20 @4 = 76. 20 раб.дней, rate 0.5.
            # НОВЫЙ: norm=80 (нет absence). total = 76+8 = 84. dev = 84-80 = +4.
            # СТАРЫЙ: отгул в absence(8); work_norm: день1 excused=min(8,4)=4->0;
            #   дни2-20=76; norm_old=76+8=84; dev = 84-84 = 0. Различает (0 vs +4).
            otgul = {1: 8.0}
            work = {d: 4.0 for d in range(2, 21)}  # 19 дней * 4 = 76
            ts = _make_timesheet(
                projects=[_make_project("25-П-1", work)],
                special_categories=[_make_category("Отгул", otgul)]
            )
            total, norm, dev = self._full_chain(ts, 0.5, _make_calendar_20wd())
            self.logger.check(
                total == 84.0 and norm == 80.0 and dev == 4.0,
                "D1: total=84, norm=80, dev=+4 (старый режим дал бы 0)",
                f"total={total}, norm={norm}, dev={dev} (ожидалось 84/80/+4)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_20_otgul_buyout_injection_D3_shortened(self):
        """§R5/D3 (ИНЪЕКЦИЯ, механизм shortened): 1.0 предпраздн.+отгул@8 -> 0/+1.

        Второй механизм кусания cap (отличный от D1): предпраздничный день,
        day_norm=7, отгул@8 > 7 -> cap min(8,7) кусает. ИНЪЕКЦИЯ: старый 0,
        новый +1.
        """
        self.logger.section("20. D3 инъекция: отгул на предпраздничном (min(8,7))")
        try:
            # rate 1.0, день20 предпраздничный. отгул@8 день20, работа дни1-19@8=152.
            # НОВЫЙ: work_norm=19*8 + 7 = 159; absence_total=0; norm=159.
            #   total = 152+8 = 160. dev = 160-159 = +1.
            # СТАРЫЙ: отгул в absence(8); день20 excused=min(8,7)=7->0; дни1-19=152;
            #   work_norm=152; norm_old=152+8=160; dev = 160-160 = 0. Различает.
            cal = _make_calendar_20wd(shortened={20})
            otgul = {20: 8.0}
            work = {d: 8.0 for d in range(1, 20)}  # 19 дней * 8 = 152
            ts = _make_timesheet(
                projects=[_make_project("25-П-1", work)],
                special_categories=[_make_category("Отгул", otgul)]
            )
            total, norm, dev = self._full_chain(ts, 1.0, cal)
            self.logger.check(
                total == 160.0 and norm == 159.0 and dev == 1.0,
                "D3: предпраздн.+отгул@8 -> dev=+1 (старый режим дал бы 0)",
                f"total={total}, norm={norm}, dev={dev} (ожидалось 160/159/+1)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_21_otgul_plus_vacation_two_classes_O4(self):
        """§R5/O4 (демо на 1.0, не различающий): два класса раздельны.

        Отпуск нейтрализует (день1 обнулён), отгул выкупает -> те же 152ч работы
        дают +8. Контроль: тот же табель БЕЗ отгула (день2=работа) -> dev 0
        (различает выкуп от нейтрализации). rate 1.0 обычный день -- ДЕМО, для
        инъекции непригоден (cap min(8,8) не кусает), но проверяет РАЗДЕЛЬНОСТЬ
        двух множеств.
        """
        self.logger.section("21. O4: отпуск нейтрализует + отгул выкупает (раздельны)")
        try:
            cal = _make_calendar_20wd()
            # отпуск@8 день1 + отгул@8 день2 + работа дни3-20@8 (18*8=144) +
            # переработка день3 +8 -> работа=152. Норма 1.0 = 160.
            # НОВЫЙ: absence_by_day={день1:8} (только отпуск); absence_total=8.
            #   work_norm: день1 excused=min(8,8)=8->0; день2 (отгул НЕ в absence)
            #   day_norm=8, excused=0 -> 8; дни3-20 = 18*8=144; work_norm=8+144=152.
            #   norm = 152 + 8 = 160. total = отпуск8 + отгул8 + работа152 = 168.
            #   dev = 168-160 = +8.
            vac = {1: 8.0}
            otgul = {2: 8.0}
            work = {d: 8.0 for d in range(3, 21)}  # дни3-20 = 144
            work[3] = 16.0  # день3: 8 работа + 8 переработка -> работа всего 152
            ts = _make_timesheet(
                projects=[_make_project("25-П-1", work)],
                special_categories=[
                    _make_category("Отпуск", vac),
                    _make_category("Отгул", otgul),
                ]
            )
            total, norm, dev = self._full_chain(ts, 1.0, cal)
            self.logger.check(
                total == 168.0 and norm == 160.0 and dev == 8.0,
                "O4: отпуск нейтрализует, отгул выкупает -> dev=+8",
                f"total={total}, norm={norm}, dev={dev} (ожидалось 168/160/+8)"
            )
            # Контроль: тот же объём работы (152ч), но БЕЗ отгула -- день2 стал
            # обычной РАБОТОЙ@8, а "переработка" день3 убрана (работа день2
            # поглотила эти 8ч, распределившись нормально). Оба варианта: 152ч
            # РЕАЛЬНОЙ работы + отпуск день1. Разница ТОЛЬКО в дне2 (отгул vs работа).
            # НОВЫЙ (без отгула): work_ctrl дни2-20@8 = 152. total = отпуск8 + 152
            #   = 160. absence_by_day={день1:8}; work_norm: день1(0)+дни2-20(152)=152;
            #   norm=152+8=160. dev = 160-160 = 0.
            # -> отгул создал +8 (выкуп), обычная работа -- 0 (нейтрализация день1).
            work_ctrl = {d: 8.0 for d in range(2, 21)}  # дни2-20 @8 = 152, без OT
            ts_ctrl = _make_timesheet(
                projects=[_make_project("25-П-1", work_ctrl)],
                special_categories=[_make_category("Отпуск", vac)]
            )
            tc, nc, dc = self._full_chain(ts_ctrl, 1.0, cal)
            self.logger.check(
                tc == 160.0 and nc == 160.0 and dc == 0.0,
                "контроль O4 (день2=работа, без отгула, 152ч работы) -> dev=0 "
                "(различает выкуп от нейтрализации)",
                f"total={tc}, norm={nc}, dev={dc} (ожидалось 160/160/0)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_22_overtime_absence_categories_skipped(self):
        """§R5 (а/б): validate_overtime пропускает absence-КАТЕГОРИИ, не ДНИ.

        (а) 0.5 полнодневный отпуск@8 -> НЕТ per-day переработки (был ложный +4).
        (б) частичный день 0.5: 4ч отпуск + 6ч работа -> переработка +2 СОХРАНЯЕТСЯ
            (пропуск категории != пропуск дня; рабочие 6ч > day_norm 4ч).
        """
        self.logger.section("22. validate_overtime: absence-категории, не дни")
        try:
            # (а) rate 0.5, отпуск@8 дни1-5 (полнодневное отсутствие) + работа
            #     дни6-20@4. Отпуск-категория пропущена -> дни1-5 без рабочих часов
            #     -> НЕТ переработки. Дни6-20 @4 = day_norm -> нет переработки.
            cal = _make_calendar_20wd()
            vac = {d: 8.0 for d in range(1, 6)}
            work = {d: 4.0 for d in range(6, 21)}
            ts_a = _make_timesheet(
                projects=[_make_project("25-П-1", work)],
                special_categories=[_make_category("Отпуск", vac)]
            )
            validator = self._make_validator(cal, rate=0.5)
            res_a = validator.validate_overtime(ts_a)
            has_overtime_a = any("Переработка" in m.message for m in res_a.messages)
            self.logger.check(
                not has_overtime_a,
                "(а) 0.5 полнодневный отпуск@8 -> НЕТ ложной per-day переработки",
                f"сообщения: {[m.message for m in res_a.messages]}"
            )

            # (б) rate 0.5, день1: 4ч отпуск + 6ч работа. Отпуск-категория
            #     пропущена -> рабочие 6ч > day_norm 4ч -> переработка +2 СОХРАНЯЕТСЯ.
            vac_b = {1: 4.0}
            work_b = {1: 6.0}
            work_b.update({d: 4.0 for d in range(2, 21)})
            ts_b = _make_timesheet(
                projects=[_make_project("25-П-1", work_b)],
                special_categories=[_make_category("Отпуск", vac_b)]
            )
            validator_b = self._make_validator(cal, rate=0.5)
            res_b = validator_b.validate_overtime(ts_b)
            has_overtime_b = any(
                "Переработка 01.01" in m.message for m in res_b.messages
            )
            self.logger.check(
                has_overtime_b,
                "(б) частичный 0.5: 4ч отпуск + 6ч работа -> переработка +2 "
                "по рабочей части СОХРАНЯЕТСЯ (пропуск категории != пропуск дня)",
                f"сообщения: {[m.message for m in res_b.messages]}"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_23_otgul_parser_gate_invisible_chars(self):
        """§R5b (ISSUE-002): отгул с zero-width/nbsp через РЕАЛЬНЫЙ parse_timesheet.

        Смена предиката _ABSENCE_NORM->_NEUTRALIZING_NORM не проверена на реальном
        парсер-пути отгула. Отгул с невидимыми символами -> корректно выкупает: в
        employee_total (факт) ЕСТЬ, в absence_by_day НЕТ (BUYOUT).
        """
        self.logger.section("23. Отгул парсер-гейт (невидимые символы, выкуп)")
        try:
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_3_parser import (
                parse_timesheet
            )
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_6_norm import (
                absence_by_day
            )
            from Daman_QGIS.utils import normalize_for_classification

            zwsp = "​"  # zero-width space
            path = os.path.join(self._tempdir, "Отгулов_01.xlsx")
            _write_timesheet_xlsx(
                path, "Отгулов Отгул Отгулович",
                rows=[
                    {"col": "D", "value": f"Отгул{zwsp}", "daily": {1: 8.0}},
                    {"col": "C", "value": "25-П-1", "daily": {2: 8.0}},
                ],
                year=2026, month=1, end_day=28
            )
            ts = parse_timesheet(path)
            self.logger.check(
                ts is not None,
                "parse_timesheet вернул табель",
                "parse_timesheet вернул None"
            )
            if ts is None:
                return
            # Отгул с zwsp распознан как категория (не проект-шифр)
            cats = {}
            for c in ts.special_categories:
                key = normalize_for_classification(c.category).lower()
                cats[key] = cats.get(key, 0.0) + sum(c.daily_hours.values())
            self.logger.check(
                cats.get("отгул", 0.0) == 8.0 and len(ts.projects) == 1,
                "Отгул<zwsp> -> категория отгул 8ч (не проект); проект 25-П-1 отдельно",
                f"cats={cats}, projects={[p.code for p in ts.projects]}"
            )
            # КЛЮЧЕВОЕ (норма-сторона): absence_by_day НЕ ловит отгул (BUYOUT
            # исключён из нормы).
            amap = absence_by_day(ts, None)
            self.logger.check(
                sum(amap.values()) == 0.0,
                "absence_by_day НЕ содержит отгул (Σ=0; отгул выкупает через факт)",
                f"Σ absence_by_day = {sum(amap.values())} (ожидалось 0.0)"
            )
            # ФАКТ-сторона (полный инвариант «отгул в employee_total ЕСТЬ»):
            # реальный _build_data_matrix даёт колонку «Отгул» с ненулём для
            # отгула с zero-width из парсер-гейта. Инвариант выкупа держится на
            # обоих концах: в факте ЕСТЬ (здесь), в норме НЕТ (проверка выше).
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_5_summary import (
                SummaryTimesheetGenerator
            )
            gen = SummaryTimesheetGenerator()
            codes, categories = gen._collect_unique_codes([ts])
            employees_data, _ = gen._build_data_matrix(
                [ts], codes, categories, end_day=None
            )
            _, hours_by_column = employees_data[0]
            # Колонка «Отгул» (каноническое имя из SPECIAL_CATEGORIES_ORDER)
            otgul_in_fact = sum(
                h for col, h in hours_by_column.items()
                if normalize_for_classification(col).lower() == "отгул"
            )
            self.logger.check(
                otgul_in_fact == 8.0,
                "employee_total СОДЕРЖИТ отгул 8ч (колонка «Отгул» _build_data_matrix)",
                f"otgul_in_fact={otgul_in_fact}, columns={list(hours_by_column.keys())} "
                f"(ожидалось 8.0)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_24_absence_placement_rejects_anomalies(self):
        """§R5d (R3b): validate_absence_placement ОТБРАКОВЫВАЕТ аномалии.

        (1) отгул на нерабочий день (day 6, выходной внутри окна) -> ошибка.
        (2) отпуск@8 + отгул@8 в один день (двойное absence) -> ошибка.
        (3) отгул@10 (Σ absence > 8) -> ошибка.
        """
        self.logger.section("24. validate_absence_placement: отбраковка аномалий")
        try:
            cal = _make_calendar_20wd()  # раб.дни 1-20; дни 21-31 нерабочие

            # (1) отгул@8 на нерабочий день 6. Календарь с "дырой": день 6 НЕ
            #     рабочий, остальные 1-20 рабочие. День 6 <= end_day(20) -> в окне
            #     проверки (не отфильтрован по cutoff).
            cal_hole = _FakeCalendar(workdays=(set(range(1, 21)) - {6}))
            ts1 = _make_timesheet(
                special_categories=[_make_category("Отгул", {6: 8.0})]
            )
            v1 = self._make_validator(cal_hole)
            r1 = v1.validate_absence_placement(ts1)
            self.logger.check(
                (not r1.is_valid) and any(
                    "нерабочий день" in m.message for m in r1.messages
                ),
                "(1) отгул на нерабочий день 6 (внутри окна) -> отбраковка",
                f"is_valid={r1.is_valid}, msgs={[m.message for m in r1.messages]}"
            )

            # (2) отпуск@8 + отгул@8 в день 3 (несовместимо)
            ts2 = _make_timesheet(
                special_categories=[
                    _make_category("Отпуск", {3: 8.0}),
                    _make_category("Отгул", {3: 8.0}),
                ]
            )
            v2 = self._make_validator(cal)
            r2 = v2.validate_absence_placement(ts2)
            self.logger.check(
                (not r2.is_valid) and any(
                    "Несовместимые" in m.message for m in r2.messages
                ),
                "(2) отпуск@8 + отгул@8 в один день -> отбраковка (несовместимо)",
                f"is_valid={r2.is_valid}, msgs={[m.message for m in r2.messages]}"
            )
            # день3 Σ absence = 16 > 8 -> ТАКЖЕ ошибка п.3 (двойная детекция ok)
            self.logger.check(
                any("превышает полный день" in m.message for m in r2.messages),
                "(2доп) Σ absence 16>8 -> также ошибка п.3",
                f"msgs={[m.message for m in r2.messages]}"
            )

            # (3) отгул@10 в день 4 (Σ absence > 8, одна категория)
            ts3 = _make_timesheet(
                special_categories=[_make_category("Отгул", {4: 10.0})]
            )
            v3 = self._make_validator(cal)
            r3 = v3.validate_absence_placement(ts3)
            self.logger.check(
                (not r3.is_valid) and any(
                    "превышает полный день" in m.message for m in r3.messages
                ),
                "(3) отгул@10 (Σ>8) -> отбраковка",
                f"is_valid={r3.is_valid}, msgs={[m.message for m in r3.messages]}"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_25_absence_placement_valid_controls(self):
        """§R5d валидные контроли (НЕ отбраковка): частичный / предпраздн.8 / отгул одиночный.

        (а) 4ч отпуск + 4ч работа в рабочий день -> ВАЛИДЕН (absence-доля 4<=8).
        (б) отпуск@8 на ПРЕДПРАЗДНИЧНОМ дне (норма 7) -> НЕ отбракован (порог 8,
            не 7; кейс G легитимен).
        (в) отгул@8 одиночный на рабочий день -> ВАЛИДЕН (выкуп легитимен).
        """
        self.logger.section("25. validate_absence_placement: валидные контроли")
        try:
            cal = _make_calendar_20wd(shortened={20})  # день20 предпраздничный

            # (а) 4ч отпуск + 4ч работа день3 -> валиден (одна absence-категория,
            #     absence-доля 4 <= 8, рабочий день).
            ts_a = _make_timesheet(
                projects=[_make_project("25-П-1", {3: 4.0})],
                special_categories=[_make_category("Отпуск", {3: 4.0})]
            )
            va = self._make_validator(cal)
            ra = va.validate_absence_placement(ts_a)
            self.logger.check(
                ra.is_valid,
                "(а) 4ч отпуск + 4ч работа (частичный) -> ВАЛИДЕН",
                f"is_valid={ra.is_valid}, msgs={[m.message for m in ra.messages]}"
            )

            # (б) отпуск@8 на предпраздничном дне 20 (норма 7) -> НЕ отбракован
            #     (порог 8, не day_norm 7; кейс G легитимен, регламент мандатирует 8).
            ts_b = _make_timesheet(
                special_categories=[_make_category("Отпуск", {20: 8.0})]
            )
            vb = self._make_validator(cal)
            rb = vb.validate_absence_placement(ts_b)
            self.logger.check(
                rb.is_valid,
                "(б) отпуск@8 на предпраздничном (порог 8, не 7) -> НЕ отбракован",
                f"is_valid={rb.is_valid}, msgs={[m.message for m in rb.messages]}"
            )

            # (в) отгул@8 одиночный на рабочий день 5 -> ВАЛИДЕН (выкуп легитимен)
            ts_c = _make_timesheet(
                special_categories=[_make_category("Отгул", {5: 8.0})]
            )
            vc = self._make_validator(cal)
            rc = vc.validate_absence_placement(ts_c)
            self.logger.check(
                rc.is_valid,
                "(в) отгул@8 одиночный на рабочий день -> ВАЛИДЕН (выкуп)",
                f"is_valid={rc.is_valid}, msgs={[m.message for m in rc.messages]}"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_26_absence_placement_calendar_runtime_failsoft(self):
        """§R5d RuntimeError (ISSUE-002): недоступный календарь -> НЕ роняет валидацию.

        F-07 graceful: п.1 (is_workday) требует сеть; при RuntimeError проверка
        размещения пропускается (валидация продолжается). п.2/п.3 (арифметика
        без сети) выполняются ДО обращения к is_workday -> двойное absence и
        Σ>8 всё равно ловятся.
        """
        self.logger.section("26. validate_absence_placement: RuntimeError fail-soft")
        try:
            cal_down = _make_calendar_20wd(raise_runtime=True)

            # (а) отгул@8 одиночный на рабочий день: п.2/п.3 чисты, п.1 нужен
            #     календарь -> недоступен -> проверка не роняет, is_valid=True.
            ts_a = _make_timesheet(
                special_categories=[_make_category("Отгул", {5: 8.0})]
            )
            va = self._make_validator(cal_down)
            ra = va.validate_absence_placement(ts_a)
            self.logger.check(
                ra.is_valid,
                "(а) недоступный календарь + валидный отгул -> НЕ роняет (fail-soft)",
                f"is_valid={ra.is_valid}, msgs={[m.message for m in ra.messages]}"
            )

            # (б) двойное absence день3 (отпуск@8+отгул@8): п.2/п.3 БЕЗ календаря
            #     ловят аномалию ДАЖЕ при недоступном календаре (is_valid=False).
            ts_b = _make_timesheet(
                special_categories=[
                    _make_category("Отпуск", {3: 8.0}),
                    _make_category("Отгул", {3: 8.0}),
                ]
            )
            vb = self._make_validator(cal_down)
            rb = vb.validate_absence_placement(ts_b)
            self.logger.check(
                (not rb.is_valid) and any(
                    "Несовместимые" in m.message for m in rb.messages
                ),
                "(б) двойное absence ловится п.2/п.3 даже без календаря",
                f"is_valid={rb.is_valid}, msgs={[m.message for m in rb.messages]}"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_27_overtime_absence_no_phantom_valid_otgul(self):
        """§R5/ISSUE-003 (стык R3/R3b): валидный одиночный отгул@8 на рабочий день.

        validate_overtime НЕ даёт ложной per-day переработки по отгул-категории
        (R3 исключил absence); validate_absence_placement НЕ бракует (одиночный
        отгул легитимен). Стык двух правок на одном валидном табеле.
        """
        self.logger.section("27. Стык R3/R3b: валидный одиночный отгул@8")
        try:
            cal = _make_calendar_20wd()
            # отгул@8 день1 + работа дни2-20@8. rate 1.0.
            ts = _make_timesheet(
                projects=[_make_project("25-П-1", {d: 8.0 for d in range(2, 21)})],
                special_categories=[_make_category("Отгул", {1: 8.0})]
            )
            validator = self._make_validator(cal, rate=1.0)

            # R3: validate_overtime -> НЕТ переработки по отгул-категории (день1)
            res_ot = validator.validate_overtime(ts)
            has_ot_day1 = any(
                "Переработка 01.01" in m.message for m in res_ot.messages
            )
            self.logger.check(
                not has_ot_day1,
                "R3: одиночный отгул@8 -> НЕТ ложной per-day переработки (день1)",
                f"сообщения: {[m.message for m in res_ot.messages]}"
            )
            # R3b: validate_absence_placement -> НЕ бракует (одиночный отгул легитимен)
            res_pl = validator.validate_absence_placement(ts)
            self.logger.check(
                res_pl.is_valid,
                "R3b: одиночный отгул@8 на рабочий день -> НЕ отбракован (легитимен)",
                f"is_valid={res_pl.is_valid}, msgs={[m.message for m in res_pl.messages]}"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")

    def test_28_validator_summary_consistency_otgul(self):
        """§R2d (D4/R5-2 распространён на выкуп): validator↔summary одно отклонение.

        На ОДНОМ табеле с отгулом (rate 0.5, Гребенников-профиль: отгул@8 +
        работа76) validator-путь (format_validation_report) и summary-путь
        (generate() -> ячейка «Отклонение») ОБЯЗАНЫ дать одинаковое отклонение
        (+4): оба считают норму через ЕДИНЫЙ helper compute_employee_norm.

        Ловит рассинхрон validator↔summary (класс R5-2/D4 ревизии 9), если бы
        кто-то развёл семантику отгула по путям. Выкуп проявляется в ЧИСЛЕ
        (+4 vs 0 на старом коде — но здесь важно РАВЕНСТВО путей, не абс.число).
        """
        self.logger.section("28. Согласованность validator↔summary (отгул выкуп)")
        try:
            import re as _re
            from Daman_QGIS.tools.F_6_special.submodules.Fsm_6_1_2_validator import (
                format_validation_report, ValidationResult
            )

            cal = _make_calendar_20wd()
            year, month, end_day = 2026, 1, 31  # весь январь; данные в днях 1-20
            fio = "Гребенников Г.Г."
            rates = {fio: 0.5}

            # rate 0.5: отгул@8 день1 + работа дни2-20@4 (=76). Выкуп: norm=80,
            # total=76+8=84, dev=+4 (D1). Табель валиден (нет аномалий/переработки).
            otgul = {1: 8.0}
            work = {d: 4.0 for d in range(2, 21)}
            ts = _make_timesheet(
                fio=fio,
                projects=[_make_project("25-П-1", work)],
                special_categories=[_make_category("Отгул", otgul)]
            )

            # --- (а) VALIDATOR-путь: format_validation_report (норма через helper) ---
            # Минимальный валидный result (норма/отклонение в отчёте не зависят
            # от result.messages -- считаются get_hours/get_employee_norm).
            valid_result = ValidationResult(is_valid=True)
            report = format_validation_report(
                [(ts, valid_result)],
                target_year=year, target_month=month,
                use_html=False, end_day=end_day,
                employee_rates=rates, calendar_manager=cal,
                calendar_unavailable=False
            )
            # Отклонение в отчёте: "Часов: <hours> (<dev>)"
            m = _re.search(r"Часов:\s*[\d.]+\s*\(([-\d.]+)\)", report)
            validator_dev = round(float(m.group(1)), 2) if m else None
            self.logger.check(
                validator_dev == 4.0,
                "validator-путь: отклонение +4 (отгул выкупил, помесячно)",
                f"validator_dev={validator_dev}; report-фрагмент: "
                f"{[l for l in report.splitlines() if 'Часов' in l]}"
            )

            # --- (б) SUMMARY-путь: generate() -> ячейка «Отклонение» ---
            report_end = date(year, month, end_day)
            path, cal_unavail = self._run_generate(
                [ts], cal, report_end, employee_rates=rates, subdir="r2d"
            )
            self.logger.check(
                path is not None and not cal_unavail,
                "generate() создал файл, календарь доступен",
                f"path={path}, cal_unavail={cal_unavail}"
            )
            if not path:
                return
            data = self._read_deviation_column(path)
            per_emp = data["per_employee"]
            summary_dev = None
            for f, v in per_emp:
                if v is not None:
                    summary_dev = round(v, 2)
                    break
            self.logger.check(
                summary_dev == 4.0,
                "summary-путь: ячейка «Отклонение» +4 (тот же helper)",
                f"summary_dev={summary_dev}, per_employee={per_emp}"
            )

            # --- КЛЮЧЕВОЕ: validator == summary (единый helper, нет рассинхрона) ---
            self.logger.check(
                validator_dev is not None and summary_dev is not None
                and validator_dev == summary_dev,
                "validator-отчёт == summary-файл: одно отклонение отгула (R2d)",
                f"validator={validator_dev}, summary={summary_dev} (обязаны совпасть)"
            )
        except Exception as e:
            self.logger.error(f"Ошибка: {e}")


def run_tests(iface, logger):
    """Точка входа для запуска тестов."""
    test = TestF61NormV2(iface, logger)
    test.run_all_tests()
    return test
